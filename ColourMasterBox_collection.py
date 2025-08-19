import sys
import os
from tkinter import SEL
import urllib.parse
import urllib.request
import pandas as pd
from PyQt5 import QtWidgets
from PyQt5 import QtCore
#from sympy import re
from form import Ui_Form
from color_form import Ui_Form as Ui_ColorDialog
from effect_form import Ui_Form as Ui_EffectDialog
from metal_form import Ui_Form as Ui_MetalDialog
from wood_form import Ui_Form as Ui_WoodDialog
import shutil
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QDialogButtonBox, QLabel, QRadioButton, QButtonGroup, QListWidgetItem, QFileDialog
from bvstd_window import Ui_Dialog 
from timsp import Ui_MainWindow
from PyQt5.QtPrintSupport import QPrintDialog, QPrinter
from PyQt5 import QtGui
import fitz  # PyMuPDF
from PyQt5.QtCore import QTranslator, QLocale, QLibraryInfo
from TCKiemtraDialog import Ui_TCDialog as Ui_TCKiemtraDialog
from PyQt5.QtGui import QIcon
from ColorwayDialog import Ui_ColorwayDialog 
from collection import Ui_Form as Ui_CollectionDialog
from defectlist import Ui_DefectListWidget
from defectinput import Ui_Form as Ui_DefectInputWidget
from printchecksheetdialog import Ui_PrintChecksheetDialog as Ui_PrintChecksheetDialog

import resources_rc

def find_company_folder():
    # List all possible company folder paths
    possible_paths = [
        r"C:\Users\Admins\The Lacquer Company\Company Files - Tài liệu",
        r"C:\Users\Admins\The Lacquer Company\Company Files - Documents",
        r"C:\Users\Admin\The Lacquer Company\Company Files - Tài liệu",
        r"C:\Users\Admin\The Lacquer Company\Company Files - Documents",
        r"C:\Users\ADMIN\The Lacquer Company\Company Files - Documents",
        r"C:\Users\ADMIN\The Lacquer Company\Company Files - Tài liệu",
        r"C:\Users\ADMINS\The Lacquer Company\Company Files - Documents",
        r"C:\Users\ADMINS\The Lacquer Company\Company Files - Tài liệu",
    ]
    for path in possible_paths:
        if os.path.exists(path):
            return path
    raise RuntimeError("Không tìm thấy thư mục công ty trên máy tính này.")

def all_words_in_text(words, text):
    return all(word in text for word in words)

class MultiSelectDialogFinal(QtWidgets.QDialog):
    def __init__(self, col4_value, matrix_ws, header_row=4, parent=None):
        super().__init__(parent)
        self.matrix_ws = matrix_ws  # <-- Add this line
        self.header_row = header_row
        self.setWindowTitle("Chọn loại kiểm tra")
        layout = QtWidgets.QVBoxLayout(self)
        msg = f"Sản phẩm {col4_value} không có trong list tên chuẩn standard."
        label = QtWidgets.QLabel(msg)
        label.setWordWrap(True)
        layout.addWidget(label)

        # Get all product names from column B, row 5 onward
        self.product_names = []
        for r in range(header_row + 1, matrix_ws.max_row + 1):
            val = matrix_ws.cell(row=r, column=2).value
            if val:
                self.product_names.append(str(val).strip())
        self.combo = QtWidgets.QComboBox()
        self.combo.addItems(self.product_names)
        layout.addWidget(QtWidgets.QLabel("Chọn sản phẩm tương tự để tự động chọn hạng mục:"))
        layout.addWidget(self.combo)

        # Checkbox options
        self.checks = []
        # Get header values from header_row
        self.header_values = []
        for col in range(1, matrix_ws.max_column + 1):
            header_val = matrix_ws.cell(row=header_row, column=col).value
            if header_val:
                self.header_values.append(str(header_val).strip())
        # Only show your desired options (Accessory, Furniture, Sơn mài, Gỗ, Metal)
        options = ["Accessory", "Furniture", "Sơn mài", "Gỗ", "Metal"]
        for opt in options:
            chk = QtWidgets.QCheckBox(opt)
            layout.addWidget(chk)
            self.checks.append(chk)

        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

        # When combobox changes, update checkboxes
        self.combo.currentIndexChanged.connect(self.update_checkboxes_for_product)
        self.update_checkboxes_for_product()  # Initial

    def update_checkboxes_for_product(self):
        idx = self.combo.currentIndex()
        if idx < 0:
            return

        row_idx = self.header_row + 1 + idx
        # For each option, check if the corresponding header column in that row has "o"
        for i, opt in enumerate(["Accessory", "Furniture", "Sơn mài", "Gỗ", "Metal"]):
            checked = False
            for col in range(1, len(self.header_values) + 1):
                header_val = self.header_values[col - 2]
                if header_val == opt:
                    cell_val = self.matrix_ws.cell(row=row_idx, column=col).value
                    if str(cell_val).strip().lower() == "o":
                        checked = True
                        break
            self.checks[i].setChecked(checked)

    def selected_options(self):
        return [chk.text() for chk in self.checks if chk.isChecked()]
    
class MultiSelectDialogCarcass(QtWidgets.QDialog):
    def __init__(self, col4_value, matrix_ws, header_row=4, parent=None):
        super().__init__(parent)
        self.matrix_ws = matrix_ws
        self.header_row = header_row
        self.setWindowTitle("Chọn loại kiểm tra (Carcass)")
        layout = QtWidgets.QVBoxLayout(self)
        msg = f"Sản phẩm {col4_value} không có trong list standard. Hãy chọn hạng mục kiểm tra."
        label = QtWidgets.QLabel(msg)
        label.setWordWrap(True)
        layout.addWidget(label)

        # Get all product names from column B, row 5 onward
        self.product_names = []
        for r in range(header_row + 1, matrix_ws.max_row + 1):
            val = matrix_ws.cell(row=r, column=2).value
            if val:
                self.product_names.append(str(val).strip())
        self.combo = QtWidgets.QComboBox()
        self.combo.addItems(self.product_names)
        layout.addWidget(QtWidgets.QLabel("Chọn sản phẩm tương tự để tự động chọn hạng mục:"))
        layout.addWidget(self.combo)

        # Checkbox options
        self.checks = []
        # Get header values from header_row
        self.header_values = []
        for col in range(1, matrix_ws.max_column + 1):
            header_val = matrix_ws.cell(row=header_row, column=col).value
            if header_val:
                self.header_values.append(str(header_val).strip())
        # Only show your desired options (Accessory, Furniture, Gỗ, Metal)
        options = ["Accessory", "Furniture", "Gỗ", "Metal"]
        for opt in options:
            chk = QtWidgets.QCheckBox(opt)
            layout.addWidget(chk)
            self.checks.append(chk)

        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

        # When combobox changes, update checkboxes
        self.combo.currentIndexChanged.connect(self.update_checkboxes_for_product)
        self.update_checkboxes_for_product()  # Initial

    def update_checkboxes_for_product(self):
        idx = self.combo.currentIndex()
        if idx < 0:
            return
        row_idx = self.header_row + 1 + idx
        for i, opt in enumerate(["Accessory", "Furniture", "Gỗ", "Metal"]):
            checked = False
            for col in range(1, len(self.header_values) + 1):
                header_val = self.header_values[col - 2]
                if header_val == opt:
                    cell_val = self.matrix_ws.cell(row=row_idx, column=col).value
                    if str(cell_val).strip().lower() == "o":
                        checked = True
                        break
            self.checks[i].setChecked(checked)

    def selected_options(self):
        return [chk.text() for chk in self.checks if chk.isChecked()]
    
class ColorSearchApp(QtWidgets.QWidget):    
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.company_folder = find_company_folder()
        self.server_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "QC FOLDER",
            "MASTER COLOR LIST QC"
        )
        serverexcel_path = os.path.join(self.server_path,"TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")

        local_path = 'TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx'
        try:
            with open(serverexcel_path, 'rb') as f:
                pass
            try:
                self.df_main = pd.read_excel(serverexcel_path, sheet_name='Lacquer FIN', header=1)
                self.df_custom = pd.read_excel(serverexcel_path, sheet_name='Custom color', header=1)
                self.df_metal = pd.read_excel(serverexcel_path, sheet_name='Metal FIN', header=1)
                self.df_wood = pd.read_excel(serverexcel_path, sheet_name='Wood FIN', header=1)
                self.df_effect = pd.read_excel(serverexcel_path, sheet_name='Effect Color Swatch Statistics', header=2)
            except Exception:
                self.df_main = pd.read_excel(local_path, sheet_name='Lacquer FIN', header=1)
                self.df_custom = pd.read_excel(local_path, sheet_name='Custom color', header=1)
                self.df_metal = pd.read_excel(local_path, sheet_name='Metal FIN', header=1)
                self.df_wood = pd.read_excel(local_path, sheet_name='Wood FIN', header=1)
                self.df_effect = pd.read_excel(local_path, sheet_name='Effect Color Swatch Statistics', header=2)

            # Strip spaces from column names for robustness
        
            self.df_main.columns = self.df_main.columns.str.strip()
            self.df_custom.columns = self.df_custom.columns.str.strip()
            self.df_metal.columns = self.df_metal.columns.str.strip()
            self.df_wood.columns = self.df_wood.columns.str.strip()
            self.df_effect.columns = self.df_effect.columns.str.strip()

            self.ui.pushButton.clicked.connect(self.search_color)
            self.ui.HieuUngButton.clicked.connect(self.search_effect_color)  # Connect the new button
            self.ui.ThemButton.clicked.connect(self.show_option_dialog)
            self.ui.BVSTDButton.clicked.connect(self.open_bvstd_window)
            self.ui.TCQCButton.clicked.connect(self.open_timsp_window)
            self.ui.collectionButton.clicked.connect(self.open_collection_dialog)
            self.ui.defectlistbutton.clicked.connect(self.open_defectlist_widget)
            self.ui.ImportExcelButton.clicked.connect(self.import_excel_dialog)

            self.ui.lineEdit.returnPressed.connect(self.search_color)  # Allow pressing Enter to search
        
            # After loading all DataFrames in __init__:
            self.all_product_names = set()
            for df in [self.df_main, self.df_custom, self.df_metal, self.df_wood]:
                if "Name" in df.columns:
                    self.all_product_names.update(df["Name"].dropna().astype(str).tolist())

            from PyQt5.QtWidgets import QCompleter

            completer = QCompleter(sorted(self.all_product_names), self)
            completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            self.ui.lineEdit.setCompleter(completer)

            # Language setup
            self.translator = QtCore.QTranslator()
            self.ui.LanguageBox.addItem(QIcon(":/images/vi.png"), "Tiếng Việt", "vi")
            self.ui.LanguageBox.addItem(QIcon(":/images/en.png"), "English", "en")
            self.ui.LanguageBox.currentIndexChanged.connect(self.change_language)
            self.ui.LanguageBox.setMinimumWidth(140)
        except PermissionError:
            QtWidgets.QMessageBox.warning(
                self,
                self.tr("Lỗi"),
                self.tr("Ai đó đang mở file excel TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx. Hãy tắt trước khi thao tác")
            )
            return
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr(f"Không thể mở file Excel:\n{e}"))


    def change_language(self):
        lang_code = self.ui.LanguageBox.currentData()
        if lang_code:
            QtWidgets.QApplication.instance().removeTranslator(self.translator)
            if self.translator.load(f"app_{lang_code}.qm"):
                QtWidgets.QApplication.instance().installTranslator(self.translator)
            # Retranslate UI
            self.ui.retranslateUi(self)

    def open_bvstd_window(self):
        self.bvstd_window = BVSTDWindow(self.company_folder, self.server_path)
        self.bvstd_window.exec_()

    def open_defectlist_widget(self):
        self.defect_widget = defectlist(self.server_path)
        self.defect_widget.show()

    def open_timsp_window(self):
        self.timsp_window = timsp(self.server_path)
        self.timsp_window.show()

    def search_color(self):
        keyword = self.ui.lineEdit.text().strip().lower()

        # Search in 'Lacquer FIN'
        words = keyword.split()
        matched = self.df_main[self.df_main['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]

        # If not found, search in 'Custom color'
        if matched.empty:
            matched = self.df_custom[self.df_custom['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
            
        if matched.empty:
            matched = self.df_metal[self.df_metal['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
        
        if matched.empty:
            matched = self.df_wood[self.df_wood['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
            if matched.empty:
                self.ui.textEdit.setText(self.tr("🔍 Không tìm thấy kết quả."))
                return

        # Add the first sentence with bold and underline
        
        html=""
        for _, row in matched.iterrows():
            name = str(row.get("Name", "")).strip()
            for col in matched.columns:
                cell = row.get(col, "")
                # If cell is a Series (shouldn't be, but just in case), get the first value
                if isinstance(cell, pd.Series):
                    cell = cell.iat[0] if not cell.empty else ""
                try:
                    is_na = pd.isna(cell)
                except Exception:
                    is_na = False
                value = "" if is_na else str(cell)
                if col.strip().lower() in ["ref image"]:
                    image_path = os.path.join(self.server_path, "Images", f"{name}.jpg")
                    if os.path.exists(image_path):
                        file_uri = urllib.parse.urljoin('file:', urllib.request.pathname2url(image_path))
                        html += f'<p><b>{col}:</b><br><img src="{file_uri}" width="200"></p>'
                    else:
                        html += f"<p><b>{col}:</b> {self.tr('Không tìm thấy ảnh')}</p>"
                else:
                    html += f"<p><b>{col}:</b> {value}</p>"
            html += "<hr>"

        self.ui.textEdit.setHtml(html)

    def search_effect_color(self):
        keyword = self.ui.lineEdit.text().strip().lower()
        # Adjust the column name below if needed
        if 'Color Name' in self.df_effect.columns:
            words = keyword.split()
            matched = self.df_effect[self.df_effect['Color Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
        else:
            self.ui.textEdit.setText(self.tr("Không tìm thấy cột 'Color Name' trong sheet hiệu ứng."))
            return

        if matched.empty:
            self.ui.textEdit.setText(self.tr("🔍 Không tìm thấy kết quả trong hiệu ứng."))
            return

        html = '<p><b><u>EFFECT COLOR</u></b></p>'
        for _, row in matched.iterrows():
            name=str(row.get("Color Name","")).strip()
            for col in matched.columns:
                cell = row.get(col, "")
                value = str(cell) if not pd.isna(cell) else ""
                if col.strip().lower() in ["ref image"]:
                    image_path=os.path.join(self.server_path,"Images",f"{name}.jpg")
                    if os.path.exists(image_path):
                        file_uri=urllib.parse.urljoin('file:',urllib.request.pathname2url(image_path))
                        html += f'<p><b>{col}:</b><br><img src="{file_uri}" width="200"></p>'
                    else:
                        html += f"<p><b>{col}:</b> ({self.tr('Không tìm thấy ảnh')})</p>"
                else:
                    html += f"<p><b>{col}:</b> {value}</p>"
            html += "<hr>"

        self.ui.textEdit.setHtml(html)

    def show_option_dialog(self):
        dialog = OptionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            selected = dialog.selected_option()
            if selected == "Color":
                color_dialog = ColorFormDialog(self)
                color_dialog.exec_()
            elif selected == "Effect":
                effect_dialog = EffectFormDialog(self)
                effect_dialog.exec_()
            elif selected == "Metal":
                metal_dialog = MetalFormDialog(self)
                metal_dialog.exec_()
            elif selected == "Wood":
                wood_dialog = WoodFormDialog(self)
                wood_dialog.exec_()
            else:
                QtWidgets.QMessageBox.information(self, self.tr("Bạn đã chọn"), f"{self.tr('Bạn đã chọn')}: {selected}")
    
    def open_collection_dialog(self):
        dialog = CollectionDialog(self.server_path)
        dialog.exec_()

    def import_excel_dialog(self):
        from PyQt5.QtWidgets import QFileDialog, QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QCheckBox, QPushButton, QHBoxLayout, QLabel
        import pandas as pd

        # Step 1: Let user select an Excel file
        file_path, _ = QFileDialog.getOpenFileName(self, "Chọn file Excel", "", "Excel Files (*.xlsx *.xls)")
        if not file_path:
            return

        # Step 2: Get all sheet names
        try:
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Lỗi", f"Không thể đọc file Excel:\n{e}")
            return

        # Step 3: Show dialog with tableView and checkboxes
        dlg = QDialog(self)
        dlg.setWindowTitle("Chọn sheet để import")
        layout = QVBoxLayout(dlg)
        label = QLabel(f"File: {file_path}")
        layout.addWidget(label)
 
        table = QTableWidget(len(sheet_names), 2)
        table.setHorizontalHeaderLabels(["Chọn", "Tên sheet"])
        table.setColumnWidth(0, 60)
        table.setColumnWidth(1, 220)
        table.verticalHeader().setVisible(False)
        # Set text alignment for all cells to center (both vertically and horizontally)
        table.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
        table.verticalHeader().setDefaultAlignment(QtCore.Qt.AlignVCenter)

        # Enable row click toggling checkbox
        def toggle_checkbox_on_row_click(index):
            if index.column() == 0: 
                # If user clicks directly on checkbox, let default behavior
                return
            row = index.row()
            chk = checkboxes[row]
            chk.setChecked(not chk.isChecked())

        table.clicked.connect(toggle_checkbox_on_row_click)
       
        checkboxes = []
        for i, sheet in enumerate(sheet_names):
            chk = QCheckBox()
            chk.setChecked(True)
            checkboxes.append(chk)
            table.setCellWidget(i, 0, chk)
            table.setItem(i, 1, QTableWidgetItem(sheet))

        layout.addWidget(table)

        btn_ok = QPushButton("OK")
        btn_cancel = QPushButton("Hủy")
        btn_unselectall = QPushButton("Unselect All")  # Add this line
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_unselectall)  # Add this line
        layout.addLayout(btn_layout)
        
        def unselect_all():
            for chk in checkboxes:
                chk.setChecked(False)
        
        btn_unselectall.clicked.connect(unselect_all)

        def on_ok():
            selected_sheets = [sheet_names[i] for i, chk in enumerate(checkboxes) if chk.isChecked()]
            if not selected_sheets:
                QtWidgets.QMessageBox.warning(dlg, "Lỗi", "Vui lòng chọn ít nhất một sheet.")
                return

            # Load all selected sheets into a dict
            try:
                xl = pd.ExcelFile(file_path)
                sheet_data = {sheet: xl.parse(sheet) for sheet in selected_sheets}
            except Exception as e:
                QtWidgets.QMessageBox.warning(dlg, "Lỗi", f"Không thể đọc sheet:\n{e}")
                return

            dlg.accept()

            # --- Show PrintChecksheetDialog ---
            class PrintChecksheetDialog(QtWidgets.QDialog):
                def __init__(self, sheet_data, file_path, server_path):
                    super().__init__()
                    self.ui = Ui_PrintChecksheetDialog()
                    self.ui.setupUi(self)
                    self.sheet_data = sheet_data
                    self.file_path = file_path
                    self.server_path = server_path
                    self.check_models = {}  # Store models for each sheet
                    self.ui.exportexcel.clicked.connect(self.export_checked_rows)
                    self.ui.lineEdit.setText(os.path.basename(file_path))

                    # Dynamically create buttons for each sheet
                    self.buttons = []
                    x, y = 19, 70
                    spacing = 10
                    font = self.font()
                    fm = QtGui.QFontMetrics(font)
                    for i, sheet in enumerate(sheet_data):
                        # Calculate width based on text
                        text_width = fm.horizontalAdvance(sheet) + 40  # 40px padding for icon/margins
                        btn_width = max(93, text_width)
                        btn = QtWidgets.QPushButton(sheet, self)
                        btn.setGeometry(x, y, btn_width, 33)
                        btn.clicked.connect(lambda checked, s=sheet: self.show_sheet(s))
                        self.buttons.append(btn)
                        x += btn_width + spacing  # Move x for next button

                    # Show the first sheet by default
                    if self.buttons:
                        self.show_sheet(self.buttons[0].text())

                    self.ui.selectallbutton.clicked.connect(self.select_all_checkboxes)
                    self.ui.unselectallbutton.clicked.connect(self.unselect_all_checkboxes)
    
                def update_checked_count(self):
                    model = self.ui.tableView.model()
                    if not model:
                        self.ui.textEdit.setText("0/0")
                        return
                    checked_count = 0
                    for row in range(model.rowCount()):
                        check_item = model.item(row, 0)
                        if check_item is not None and check_item.checkState() == QtCore.Qt.Checked:
                            checked_count += 1
                    total_count = model.rowCount()
                    self.ui.textEdit.setText(f"{checked_count}/{total_count}")
                    self.ui.textEdit.setAlignment(QtCore.Qt.AlignCenter)  # Center text horizontally and vertically
    
                def show_sheet(self, sheet_name):
                    df = self.sheet_data[sheet_name]
                    # Find the row where column A is "NO" (case-insensitive)
                    header_row = None
                    for idx, val in enumerate(df.iloc[:, 0]):
                        if str(val).strip().upper() == "NO":
                            header_row = idx
                            break
                    if header_row is None:
                        QtWidgets.QMessageBox.warning(self, "Lỗi", "Không tìm thấy dòng tiêu đề (NO) trong sheet.")
                        return

                    # Set header from that row, and data from the next row
                    headers = ["Chọn"] + [str(col) for col in df.iloc[header_row]]
                    df_content = df.iloc[header_row + 1:].copy()

                    # --- Clean data ---
                    import re
                    # Remove extra spaces in all string cells
                    df_content = df_content.applymap(
                        lambda x: re.sub(r'\s{2,}', ' ', str(x).strip()) if isinstance(x, str) else x
                    )
                    # Drop rows where column 3 (index 2) is blank or only spaces
                    df_content = df_content[df_content.iloc[:, 2].astype(str).str.strip() != "nan"]

                    # --- Continue with your existing code ---
                    model = QtGui.QStandardItemModel()
                    model.setColumnCount(len(headers))
                    model.setHorizontalHeaderLabels(headers)
                    for _, row in df_content.iterrows():
                        items = []
                        check_item = QtGui.QStandardItem()
                        check_item.setCheckable(True)
                        check_item.setCheckState(QtCore.Qt.Unchecked)
                        items.append(check_item)
                        for value in row:
                            item = QtGui.QStandardItem(str(value))
                            items.append(item)
                        model.appendRow(items)
                    self.ui.tableView.setModel(model)
                    self.ui.tableView.setColumnWidth(0, 50)
                    self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
                    self.check_models[sheet_name] = model
                    self.last_shown_sheet = sheet_name
                    # Add this to enable row click toggling checkbox
                    def toggle_checkbox_on_row_click1(index):
                        if index.column() == 0:
                            # If user clicks directly on checkbox, let default behavior
                            return
                        row = index.row()
                        check_item = model.item(row, 0)
                        if check_item is not None:
                            if check_item.checkState() == QtCore.Qt.Checked:
                                check_item.setCheckState(QtCore.Qt.Unchecked)
                            else:
                                check_item.setCheckState(QtCore.Qt.Checked)
                            self.update_checked_count()
                    self.ui.tableView.clicked.connect(toggle_checkbox_on_row_click1)
    
                    # Also connect each checkbox to update_checked_count
                    for row in range(model.rowCount()):
                        check_item = model.item(row, 0)
                        if check_item is not None:
                            check_item.itemChanged = lambda: self.update_checked_count()

                    self.update_checked_count()  # Initial update

                def unselect_all_checkboxes(self):
                    model = self.ui.tableView.model()
                    if not model:
                        return
                    checked_count = 0
                    for row in range(model.rowCount()):
                        check_item = model.item(row, 0)
                        if check_item is not None:
                            check_item.setCheckState(QtCore.Qt.Unchecked)
                    total_count = model.rowCount()
                    self.ui.textEdit.setText(f"{checked_count}/{total_count}")
                    self.ui.textEdit.setAlignment(QtCore.Qt.AlignCenter)  # Center text horizontally and vertically                

                def select_all_checkboxes(self):
                    model = self.ui.tableView.model()
                    if not model:
                        return
                    checked_count = 0
                    for row in range(model.rowCount()):
                        check_item = model.item(row, 0)
                        if check_item is not None:
                            check_item.setCheckState(QtCore.Qt.Checked)
                            checked_count += 1
                    total_count = model.rowCount()
                    self.ui.textEdit.setText(f"{checked_count}/{total_count}")
                    self.ui.textEdit.setAlignment(QtCore.Qt.AlignCenter)  # Center text horizontally and vertically

                def export_checked_rows(self):
                    from openpyxl.styles import Border, Side
                    def add_bottom_border(ws, last_row, start_col=1, end_col=16):
                        thin = Side(border_style="thin", color="000000")
                        for col in range(start_col, end_col + 1):
                            cell = ws.cell(row=last_row, column=col)
                            # Keep existing borders for other sides
                            cell.border = Border(
                                left=cell.border.left,
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=thin
                            )
                            
                    import openpyxl
                    from PyQt5.QtWidgets import QFileDialog, QMessageBox
                    import re, os

                   # Get the currently selected sheet button
                    current_sheet = None
                    for btn in self.buttons:
                        if btn.isChecked():
                            current_sheet = btn.text()
                            break
                    if not current_sheet:
                        # Fallback: use the last shown sheet
                        current_sheet = self.last_shown_sheet if hasattr(self, "last_shown_sheet") else self.buttons[0].text()

                    model = self.ui.tableView.model()
                    if not model or model.rowCount() == 0:
                        QMessageBox.warning(self, "Lỗi", "Không có dữ liệu để xuất.")
                        return

                    # Collect checked rows
                    checked_rows = []
                    for row in range(model.rowCount()):
                        if model.item(row, 0).checkState() == QtCore.Qt.Checked:
                            checked_rows.append(row)
                    if not checked_rows:
                        QMessageBox.warning(self, "Lỗi", "Vui lòng chọn ít nhất một dòng để xuất.")
                        return

                    # Find PO number (e.g. PO4, PO123, etc.) anywhere in the filename
                    po_number = "PO_UNKNOWN"
                    match = re.search(r'PO\d+[-_]\d+', os.path.basename(self.file_path), re.IGNORECASE)
                    if match:
                        po_number = match.group(0)

                    # Ask user for a folder to save all files
                    folder = QFileDialog.getExistingDirectory(self, "Chọn thư mục để lưu các file Excel")
                    if not folder:
                        return

                    template_path = os.path.join(self.server_path, "QC Check list - Format.xlsx")
                    template_path2 = os.path.join(self.server_path, "QC Check list - Format carcass - vi.xlsx")
                    if not os.path.exists(template_path) or not os.path.exists(template_path2):
                        QMessageBox.warning(self, "Lỗi", "Không tìm thấy file format.")
                        return

                    # Path to List Product QC.xlsx
                    qc_path = os.path.join(self.server_path, "List Product QC.xlsx")
                    if not os.path.exists(qc_path):
                        QMessageBox.warning(self, "Lỗi", f"Không tìm thấy file: {qc_path}")
                        return

                    success_count_final = 0
                    success_count_carcass = 0
                    for row in checked_rows:
                        col4_value = model.item(row, 4).text() if model.item(row, 4) else ""
                        col2 = model.item(row, 2).text() if model.item(row, 2) else ""
                        col3 = model.item(row, 3).text() if model.item(row, 3) else ""
                        col5 = model.item(row, 5).text() if model.item(row, 5) else ""
                        col6 = 0  # Start as a number
                        for r in range(model.rowCount()):
                            val2 = model.item(r, 2).text() if model.item(r, 2) else ""
                            val3 = model.item(r, 3).text() if model.item(r, 3) else ""
                            val4 = model.item(r, 4).text() if model.item(r, 4) else ""
                            if val2 == col2 and val3 == col3 and val4 == col4_value:
                                val6 = model.item(r, 6).text() if model.item(r, 6) else "0"
                                try:
                                    col6 += float(val6)
                                except ValueError:
                                    pass  # skip if not a number

                        # Remove "CUSTOM" from col4_value for searching
                        import re
                        col4_value_clean = re.sub(r'\s*CUSTOM\s*', '', col4_value, flags=re.IGNORECASE).strip()

                        save_filename = f"{po_number}_{col2}-{col3}_Final_{current_sheet}_{col4_value}.xlsx"
                        save_filename = re.sub(r'[\\/*?:"<>|]', "_", save_filename)
                        save_path = os.path.join(folder, save_filename)
                        if col5=="Carcass"or col5=="CARCASS":
                            save_filename2 = f"{po_number}_{col2}-{col3}_OnlyCarcass_{current_sheet}_{col4_value}.xlsx"
                        else:
                            save_filename2 = f"{po_number}_{col2}-{col3}_Carcass+Final_{current_sheet}_{col4_value}.xlsx"
                        save_filename2 = re.sub(r'[\\/*?:"<>|]', "_", save_filename2)
                        save_path2 = os.path.join(folder, save_filename2)

                        # Open both templates
                        wb = openpyxl.load_workbook(template_path)
                        wb2 = openpyxl.load_workbook(template_path2)
                        ws = wb["Format_fin_vi"]
                        ws2 = wb2["Format_car_vi"]

                        # Write common info to both sheets
                        for ws_target in [ws, ws2]:
                            ws_target["C4"].value = po_number
                            ws_target["G4"].value = current_sheet
                            ws_target["L4"].value = col4_value
                            ws_target["G5"].value = f"{col2}-{col3}".strip()
                            ws_target["C5"].value = col5
                        ws["L5"].value = col6
                        ws2["L5"].value = 1  # Default to 1 for carcass
                        
                        # Open QC file
                        qc_wb = openpyxl.load_workbook(qc_path, data_only=True)

                        # --- Format_fin_vi logic ---
                        if "ANH THIỆP" in current_sheet.upper():
                            matrix_ws = qc_wb["MatrixATHIEP"]
                        else:
                            matrix_ws = qc_wb["Matrix"]
                        hangmuc_ws = qc_wb["Hangmuc"]
                        matrix_row_idx = None
                        search_names = [str(col4_value_clean)]
                        # Set print area for ws
                        def last_data_row(ws, max_col=12):
                            for row in range(ws.max_row, 0, -1):
                                for col in range(1, max_col + 1):
                                    if ws.cell(row=row, column=col).value not in (None, ""):
                                        return row
                            return 1
                        
                        # 1. Find row in Matrix where column B == col4_value_clean (from row 5)
                        if col5 == "Carcass" or col5 == "CARCASS":
                            pass  # Skip if col5 is Carcass, handled separately below
                        else:
                            if not str(col4_value_clean).upper().endswith("TABLE"):
                                search_names.append(str(col4_value_clean).strip() + " TABLE")
                            for r in range(5, matrix_ws.max_row + 1):
                                cell_val = matrix_ws.cell(row=r, column=2).value
                                cell_val_str = str(cell_val).strip() if cell_val else ""
                                for name in search_names:
                                    if cell_val_str.upper() == name.upper():
                                        matrix_row_idx = r
                                        break
                                if matrix_row_idx is not None:
                                    break
                            if matrix_row_idx is None:
                                # Show multi-select dialog
                                dlg = MultiSelectDialogFinal(col4_value, matrix_ws, header_row=4)
                                if dlg.exec_() == QtWidgets.QDialog.Accepted:
                                    selected_headers = dlg.selected_options()
                                    if not selected_headers:
                                        QtWidgets.QMessageBox.warning(self, "Lỗi", "Bạn chưa chọn loại kiểm tra nào.")
                                        continue
                                    o_columns = selected_headers
                                else:
                                    continue  # User cancelled
                            else:
                                o_columns = []
                                for col in range(1, matrix_ws.max_column + 1):
                                    val = matrix_ws.cell(row=matrix_row_idx, column=col).value
                                    if str(val).strip().lower() == "o":
                                        header_val = matrix_ws.cell(row=4, column=col).value
                                        if header_val:
                                            o_columns.append(header_val)
                            hangmuc_rows = []
                            for header in o_columns:
                                for r in range(1, hangmuc_ws.max_row + 1):
                                    if str(hangmuc_ws.cell(row=r, column=1).value).strip() == str(header).strip():
                                        hangmuc_rows.append([
                                            hangmuc_ws.cell(row=r, column=1).value,  # A
                                            hangmuc_ws.cell(row=r, column=2).value,  # B
                                            hangmuc_ws.cell(row=r, column=3).value,  # C
                                            hangmuc_ws.cell(row=r, column=4).value,  # D
                                            hangmuc_ws.cell(row=r, column=5).value   # E
                                        ])
                            for i, hm_row in enumerate(hangmuc_rows):
                                ws.cell(row=9 + i, column=2).value = hm_row[0]  # A
                                ws.cell(row=9 + i, column=1).value = hm_row[1]  # B
                                ws.cell(row=9 + i, column=4).value = hm_row[2]  # D
                                ws.cell(row=9 + i, column=7).value = hm_row[3]  # G
                                ws.cell(row=9 + i, column=12).value = hm_row[4] # L

                            last_row = last_data_row(ws, max_col=12)
                            ws.print_area = f"A1:P{last_row}"
                            add_bottom_border(ws, last_row, start_col=1, end_col=16)

                            ws.oddHeader.center.text = f"{po_number}_{col2}-{col3}_Final_{current_sheet}_{col4_value}"
                            from datetime import datetime
                            ws.oddHeader.right.text = datetime.today().strftime("Date: %d/%m/%Y")

                        # --- Format_car_vi logic (same, but for MatrixCarcass/HangmucCarcass and ws2) ---
                        if "MatrixCarcass" in qc_wb.sheetnames and "HangmucCarcass" in qc_wb.sheetnames:
                            if "ANH THIỆP" in current_sheet.upper():
                                matrix_ws2 = qc_wb["MatrixCarcassATHIEP"]
                            else:
                                matrix_ws2 = qc_wb["MatrixCarcass"]
                            hangmuc_ws2 = qc_wb["HangmucCarcass"]
                            # Find row in MatrixCarcass where column B == col4_value_clean (from row 5)
                            matrix_row_idx2 = None
                            for r in range(5, matrix_ws2.max_row + 1):
                                cell_val = matrix_ws2.cell(row=r, column=2).value
                                if str(cell_val).strip() == str(col4_value_clean):
                                    matrix_row_idx2 = r
                                    break
                            matrix_row_idx2 = None
                            for r in range(5, matrix_ws2.max_row + 1):
                                cell_val = matrix_ws2.cell(row=r, column=2).value
                                cell_val_str = str(cell_val).strip() if cell_val else ""
                                for name in search_names:
                                    if cell_val_str.upper() == name.upper():
                                        matrix_row_idx2 = r
                                        break
                                if matrix_row_idx2 is not None:
                                    break
                            if matrix_row_idx2 is None:
                                # Show multi-select dialog
                                dlg1 = MultiSelectDialogCarcass(col4_value, matrix_ws2, header_row=4)
                                if dlg1.exec_() == QtWidgets.QDialog.Accepted:
                                    selected_headers = dlg1.selected_options()
                                    if not selected_headers:
                                        QtWidgets.QMessageBox.warning(self, "Lỗi", "Bạn chưa chọn loại kiểm tra nào.")
                                        continue
                                    o_columns2 = selected_headers
                                else:
                                    continue  # User cancelled
                            else:
                                o_columns2 = []
                                for col in range(1, matrix_ws2.max_column + 1):
                                    val = matrix_ws2.cell(row=matrix_row_idx2, column=col).value
                                    if str(val).strip().lower() == "o":
                                        header_val = matrix_ws2.cell(row=4, column=col).value
                                        if header_val:
                                            o_columns2.append(header_val)
                            hangmuc_rows2 = []
                            for header in o_columns2:
                                for r in range(1, hangmuc_ws2.max_row + 1):
                                    if str(hangmuc_ws2.cell(row=r, column=1).value).strip() == str(header).strip():
                                        hangmuc_rows2.append([
                                            hangmuc_ws2.cell(row=r, column=1).value,  # A
                                            hangmuc_ws2.cell(row=r, column=2).value,  # B
                                            hangmuc_ws2.cell(row=r, column=3).value,  # C
                                            hangmuc_ws2.cell(row=r, column=4).value,  # D
                                            hangmuc_ws2.cell(row=r, column=5).value   # E
                                        ])
                            for i, hm_row in enumerate(hangmuc_rows2):
                                ws2.cell(row=9 + i, column=2).value = hm_row[0]  # A
                                ws2.cell(row=9 + i, column=1).value = hm_row[1]  # B
                                ws2.cell(row=9 + i, column=4).value = hm_row[2]  # D
                                ws2.cell(row=9 + i, column=7).value = hm_row[3]  # G
                                ws2.cell(row=9 + i, column=12).value = hm_row[4] # L

                            # Set print area for ws2
                            last_row2 = last_data_row(ws2, max_col=12)
                            ws2.print_area = f"A1:P{last_row2}"
                            ws2.oddHeader.center.text = f"{po_number}_{col2}-{col3}_Carcass_{current_sheet}_{col4_value}"
                            from datetime import datetime
                            ws2.oddHeader.right.text = datetime.today().strftime("Date: %d/%m/%Y")

                        # Save the combined workbook
                        try:
                            if col5 == "Carcass" or col5 == "CARCASS":
                                wb2.save(save_path2)
                                success_count_carcass += 1
                            else:
                                ws2.oddHeader.center.text = f"{po_number}_{col2}-{col3}_Carcass+Final_{current_sheet}_{col4_value}"
                                wb.save(save_path)
                                wb2.save(save_path2)
                                success_count_final += 1
                                success_count_carcass += 1

                                # --- Combine wb and wb2 (A to P, all rows, keep format/merge) ---
                                import xlwings as xw
                                with xw.App(visible=False) as app:
                                    app.display_alerts = False
                                    app.screen_updating = False

                                    wb_xlw = app.books.open(save_path)
                                    wb2_xlw = app.books.open(save_path2)

                                    ws = wb_xlw.sheets[0]
                                    ws2 = wb2_xlw.sheets[0]

                                    # Find last used row in ws (final file)
                                    last_row = ws.range("A" + str(ws.cells.last_cell.row)).end("up").row

                                    # Find last used row in ws2 (carcass file)
                                    last_row2 = ws2.range("A" + str(ws2.cells.last_cell.row)).end("up").row

                                    # Copy A1:P{last_row} from ws
                                    rng_to_copy = ws.range(f"A2:P{last_row}")
                                    # Paste to ws2, starting at first empty row (last_row2 + 2)
                                    dest_rng = ws2.range(f"A{last_row2 + 2}")
                                    rng_to_copy.api.Copy(dest_rng.api)

                                    # Set row height for the first pasted row (last_row2 + 1)
                                    ws2.range(f"{last_row2 + 1}:{last_row2 + 1}").row_height = 27

                                    # Find new last row after paste
                                    new_last_row2 = ws2.range("A" + str(ws2.cells.last_cell.row)).end("up").row

                                    # Set print area again (A1:P{new_last_row2})
                                    ws2.api.PageSetup.PrintArea = f"$A$1:$P${new_last_row2}"

                                    wb2_xlw.save()
                                    wb2_xlw.close()
                                    wb_xlw.close()
                                    if os.path.exists(save_path):
                                        try:
                                            os.remove(save_path)
                                        except Exception as e:
                                            QtWidgets.QMessageBox.warning(self, "Lỗi", f"Không thể xóa file tạm thời:\n{e}")
                                    
                                    
                        except Exception as e:
                            QMessageBox.critical(self, "Lỗi", f"Không thể lưu file Excel:\n{e}")                            

                    QMessageBox.information(self, "Thành công", f"Đã xuất {success_count_final} file Excel Final và {success_count_carcass} file Excel Carcass thành công!")
            # Show the dialog
            print_dialog = PrintChecksheetDialog(sheet_data, file_path, self.server_path)
            print_dialog.setWindowTitle("Print Checksheet")
            print_dialog.exec_()

        btn_ok.clicked.connect(on_ok)
        btn_cancel.clicked.connect(dlg.reject)

        dlg.exec_()

class OptionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Choose an Option")
        self.resize(280, 200)
        self.setMinimumWidth(280)
        layout = QVBoxLayout(self)

        label = QLabel("Choose a type:")
        layout.addWidget(label)

        self.button_group = QButtonGroup(self)
        self.radio_color = QRadioButton("Color")
        self.radio_metal = QRadioButton("Metal")
        self.radio_wood = QRadioButton("Wood")
        self.radio_effect = QRadioButton("Effect")
        self.button_group.addButton(self.radio_color)
        self.button_group.addButton(self.radio_metal)
        self.button_group.addButton(self.radio_wood)
        self.button_group.addButton(self.radio_effect)
        layout.addWidget(self.radio_color)
        layout.addWidget(self.radio_metal)
        layout.addWidget(self.radio_wood)
        layout.addWidget(self.radio_effect)

        self.radio_color.setChecked(True)  # Default selection

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_option(self):
        if self.radio_color.isChecked():
            return "Color"
        elif self.radio_metal.isChecked():
            return "Metal"
        elif self.radio_wood.isChecked():
            return "Wood"
        elif self.radio_effect.isChecked():
            return "Effect"
        return None

class ColorFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_ColorDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        
        import pandas as pd
        from PyQt5.QtWidgets import QCompleter

        # Load product names from "Lacquer FIN"
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        product_names = []
        try:
            df = pd.read_excel(excel_path, sheet_name="Lacquer FIN", header=1)
            if "Name" in df.columns:
                product_names = df["Name"].dropna().astype(str).tolist()
        except Exception as e:
            product_names = []

        completer = QCompleter(sorted(product_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit_3.setCompleter(completer)
        
        
        self.them_mau_button = self.ui.ThemMauButton
        self.them_mau_button.clicked.connect(self.add_new_color)
        self.up_anh_button = self.ui.UpAnhButton
        self.up_anh_button.clicked.connect(self.upload_image)
        self.uploaded_image_path = None  # <-- Add this line
        self.ui.CapnhatMauButton.clicked.connect(self.update_color)

    def add_new_color(self):
        # Map lineEdits to columns
        mapping = [
            ("lineEdit_1", "Collection"),
            ("lineEdit_2", "Ref-Tone Code"),
            ("lineEdit_3", "Name"),
            ("lineEdit_4", "Reference"),
            ("lineEdit_5", "Status"),
            ("lineEdit_6", "Generation"),
            ("lineEdit_7", "Process"),
            ("lineEdit_8", "Request day"),
            ("lineEdit_9", "Qty"),
            ("lineEdit_10", "Approved by "),
            ("lineEdit_11", "Approved day"),
            ("lineEdit_12", "Sup - incharge"),
            ("lineEdit_13", "Master"),
            ("lineEdit_14", "Notes"),
            ("lineEdit_15", "Applied"),
        ]

        # Read values from lineEdits
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        # Load Excel file
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Lacquer FIN",header=1)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        # Append new row
        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        # Save back to Excel
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Lacquer FIN", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from "Name" as the filename
                image_name = row_data.get("Name", "new_image").strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã thêm màu mới vào danh sách."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr("Chọn ảnh JPG"),
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path  # <-- Save the path

    def update_color(self):
        # Mapping of line edits to DataFrame columns
        mapping = {
            "lineEdit_1": "Collection",
            "lineEdit_2": "Ref-Tone Code",
            "lineEdit_3": "Name",
            "lineEdit_4": "Reference",
            "lineEdit_5": "Status",
            "lineEdit_6": "Generation",
            "lineEdit_7": "Process",
            "lineEdit_8": "Request day",
            "lineEdit_9": "Qty",
            "lineEdit_10": "Approved by ",
            "lineEdit_11": "Approved day",
            "lineEdit_12": "Sup - incharge",
            "lineEdit_13": "Master",
            "lineEdit_14": "Notes",
            "lineEdit_15": "Applied",
        }

        # Read the current values from the line edits
        row_data = {}
        for edit_name, col_name in mapping.items():
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        # Validate the "Name" field
        if not row_data["Name"]:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Tên màu không được để trống."))
            return

        # Load the existing data
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Lacquer FIN", header=1)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        # Find the row to update
        try:
            row_index = df[df["Name"] == row_data["Name"]].index[0]
        except IndexError:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Không tìm thấy màu để cập nhật."))
            return

        # Update the row
        # Only update non-empty fields
        for col, value in row_data.items():
            if value:  # Only update if not empty
                df.at[row_index, col] = value

        # Save the changes
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Lacquer FIN", index=False)

            # Copy the new image if provided
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                image_name = row_data["Name"].strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã cập nhật màu."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

class EffectFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_EffectDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        self.uploaded_image_path = None
        import pandas as pd
        from PyQt5.QtWidgets import QCompleter

        # Load product names from "Effect Color Swatch Statistics"
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        effect_names = []
        try:
            df = pd.read_excel(excel_path, sheet_name="Effect Color Swatch Statistics", header=2)
            if "Color Name" in df.columns:
                effect_names = df["Color Name"].dropna().astype(str).tolist()
        except Exception:
            effect_names = []

        completer = QCompleter(sorted(effect_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit_1.setCompleter(completer)

        self.ui.UpAnhHUButton.clicked.connect(self.upload_image)
        self.ui.ThemHUButton.clicked.connect(self.add_new_effect)
        self.ui.CapnhatHUButton.clicked.connect(self.update_effect)

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr("Chọn ảnh JPG"),
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path

    def add_new_effect(self):
        mapping = [
            ("lineEdit_1", "Color Name"),
            ("lineEdit_2", "Qty"),
            ("lineEdit_3", "Approval date"),
            ("lineEdit_4", "Note"),
        ]
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Effect Color Swatch Statistics", header=2)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Effect Color Swatch Statistics", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from lineEdit_1 as the filename
                image_name = self.ui.lineEdit_1.text().strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã thêm hiệu ứng mới vào danh sách."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

    def update_effect(self):
        mapping = {
            "lineEdit_1": "Color Name",
            "lineEdit_2": "Qty",
            "lineEdit_3": "Approval date",
            "lineEdit_4": "Note",
        }

        # Read the current values from the line edits
        row_data = {}
        for edit_name, col_name in mapping.items():
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        # Validate the "Color Name" field
        if not row_data["Color Name"]:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Tên màu không được để trống."))
            return

        # Load the existing data
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Effect Color Swatch Statistics", header=2)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        # Find the row to update
        try:
            row_index = df[df["Color Name"] == row_data["Color Name"]].index[0]
        except IndexError:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Không tìm thấy hiệu ứng để cập nhật."))
            return

        # Update the row
        # Only update non-empty fields
        for col, value in row_data.items():
            if value:  # Only update if not empty
                df.at[row_index, col] = value

        # Save the changes
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Effect Color Swatch Statistics", index=False)

            # Copy the new image if provided
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                image_name = row_data["Color Name"].strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã cập nhật hiệu ứng."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

class MetalFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_MetalDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        self.uploaded_image_path = None

        import pandas as pd
        from PyQt5.QtWidgets import QCompleter

        # Load metal names from "Metal FIN"
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        metal_names = []
        try:
            df = pd.read_excel(excel_path, sheet_name="Metal FIN", header=1)
            if "Name" in df.columns:
                metal_names = df["Name"].dropna().astype(str).tolist()
        except Exception:
            metal_names = []

        completer = QCompleter(sorted(metal_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit_2.setCompleter(completer)
        self.ui.UpAnhMeButton.clicked.connect(self.upload_image)
        self.ui.ThemMetalButton.clicked.connect(self.add_new_metal)
        self.ui.CapnhatMetalButton.clicked.connect(self.update_metal)

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr("Chọn ảnh JPG"),
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path

    def add_new_metal(self):
        mapping = [
            ("lineEdit_1", "Code"),
            ("lineEdit_2", "Name"),
            ("lineEdit_3", "Status"),
            ("lineEdit_4", "Description"),
            ("lineEdit_5", "Generation"),
            ("lineEdit_6", "Process"),
            ("lineEdit_7", "Request day"),
            ("lineEdit_8", "Qty"),
            ("lineEdit_9", "Approved by"),
            ("lineEdit_10", "Approved day"),
            ("lineEdit_11", "Reject"),
            ("lineEdit_12", "Actual end day"),
            ("lineEdit_13", "Supplier"),
            ("lineEdit_14", "Notes"),
            ("lineEdit_15", "Reference"),
        ]
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Metal FIN", header=1)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Metal FIN", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from lineEdit_2 ("Name") as the filename
                image_name = self.ui.lineEdit_2.text().strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã thêm kim loại mới vào danh sách."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

    def update_metal(self):
        # Mapping of line edits to DataFrame columns
        mapping = {
            "lineEdit_1": "Code",
            "lineEdit_2": "Name",
            "lineEdit_3": "Status",
            "lineEdit_4": "Description",
            "lineEdit_5": "Generation",
            "lineEdit_6": "Process",
            "lineEdit_7": "Request day",
            "lineEdit_8": "Qty",
            "lineEdit_9": "Approved by",
            "lineEdit_10": "Approved day",
            "lineEdit_11": "Reject",
            "lineEdit_12": "Actual end day",
            "lineEdit_13": "Supplier",
            "lineEdit_14": "Notes",
            "lineEdit_15": "Reference",
        }

        # Read the current values from the line edits
        row_data = {}
        for edit_name, col_name in mapping.items():
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text().strip()

        # Validate the "Name" field (from lineEdit_2)
        if not row_data["Name"]:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Tên kim loại không được để trống."))
            return

        # Load the existing data
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Metal FIN", header=1)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        # Find the row to update by "Name"
        try:
            row_index = df[df["Name"] == row_data["Name"]].index[0]
        except IndexError:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Không tìm thấy kim loại để cập nhật."))
            return

        # Only update non-empty fields (except "Name")
        for edit_name, col_name in mapping.items():
            if col_name == "Name":
                continue
            value = row_data[col_name]
            if value:
                df.at[row_index, col_name] = value

        # Save the changes
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Metal FIN", index=False)

            # Copy the new image if provided
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                image_name = row_data["Name"].strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã cập nhật kim loại."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

class WoodFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_WoodDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        self.uploaded_image_path = None

        import pandas as pd
        from PyQt5.QtWidgets import QCompleter

        # Load wood names from "Wood FIN"
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        wood_names = []
        try:
            df = pd.read_excel(excel_path, sheet_name="Wood FIN", header=1)
            if "Name" in df.columns:
                wood_names = df["Name"].dropna().astype(str).tolist()
        except Exception:
            wood_names = []

        completer = QCompleter(sorted(wood_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit_2.setCompleter(completer)
        self.ui.UpAnhGoButton.clicked.connect(self.upload_image)
        self.ui.ThemGoButton.clicked.connect(self.add_new_wood)
        self.ui.CapnhatGoButton.clicked.connect(self.update_wood)

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr("Chọn ảnh JPG"),
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path

    def add_new_wood(self):
        mapping = [
            ("lineEdit_1", "Code"),
            ("lineEdit_2", "Name"),
            ("lineEdit_3", "Status"),
            ("lineEdit_4", "PO"),
            ("lineEdit_5", "Generation"),
            ("lineEdit_6", "Process"),
            ("lineEdit_7", "Request day"),
            ("lineEdit_8", "Qty"),
            ("lineEdit_9", "Approved by"),
            ("lineEdit_10", "Approved day"),
            ("lineEdit_11", "Reject"),
            ("lineEdit_12", "Actual end day"),
            ("lineEdit_13", "Supplier"),
            ("lineEdit_14", "Notes"),
            ("lineEdit_15", "Master"),
            ("lineEdit_16", "Reference"),
        ]
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Wood FIN", header=1)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Wood FIN", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from lineEdit_2 ("Name") as the filename
                image_name = self.ui.lineEdit_2.text().strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã thêm gỗ mới vào danh sách."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

    def update_wood(self):
        # Mapping of line edits to DataFrame columns
        mapping = {
            "lineEdit_1": "Code",
            "lineEdit_2": "Name",
            "lineEdit_3": "Status",
            "lineEdit_4": "PO",
            "lineEdit_5": "Generation",
            "lineEdit_6": "Process",
            "lineEdit_7": "Request day",
            "lineEdit_8": "Qty",
            "lineEdit_9": "Approved by",
            "lineEdit_10": "Approved day",
            "lineEdit_11": "Reject",
            "lineEdit_12": "Actual end day",
            "lineEdit_13": "Supplier",
            "lineEdit_14": "Notes",
            "lineEdit_15": "Master",
            "lineEdit_16": "Reference",
        }

        # Read the current values from the line edits
        row_data = {}
        for edit_name, col_name in mapping.items():
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text().strip()

        # Validate the "Name" field (from lineEdit_2)
        if not row_data["Name"]:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Tên gỗ không được để trống."))
            return

        # Load the existing data
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Wood FIN", header=1)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        # Find the row to update by "Name"
        try:
            row_index = df[df["Name"] == row_data["Name"]].index[0]
        except IndexError:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Không tìm thấy gỗ để cập nhật."))
            return

        # Only update non-empty fields (except "Name")
        for edit_name, col_name in mapping.items():
            if col_name == "Name":
                continue
            value = row_data[col_name]
            if value:
                df.at[row_index, col_name] = value

        # Save the changes
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Wood FIN", index=False)

            # Copy the new image if provided
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                image_name = row_data["Name"].strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã cập nhật gỗ."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

class BVSTDWindow(QDialog):
    def __init__(self, company_folder, server_path):
        super().__init__()
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

        self.company_folder = company_folder
        self.server_path = server_path 

        self.target_folder = os.path.join(company_folder, "TLC DRAWINGS", "STANDARD DRAWNGS  SEND TO SUPPLIER")

        # Sự kiện
        self.ui.searchLineEdit.textChanged.connect(self.perform_search)
        self.ui.resultList.itemClicked.connect(self.show_preview)
        self.ui.downloadButton.clicked.connect(self.download_file)
        self.ui.PrintButton.clicked.connect(self.print_image)
        self.ui.ColorwayButton.clicked.connect(self.open_colorway_ppt)
        self.ui.OpenButton.clicked.connect(self.open_selected_file)

        self.found_files = []

    def perform_search(self):
        keyword = self.ui.searchLineEdit.text().strip().lower()
        if not keyword:
            self.ui.resultList.clear()
            self.found_files = []
            return

        # Build all search patterns
        patterns = [
            f"standard-{keyword}",
            f"standard- {keyword}",
            f"standard - {keyword}",
            f"standard -{keyword}",
            keyword
        ]

        self.ui.resultList.clear()
        self.found_files = []

        if not os.path.exists(self.target_folder):
            return

        for root, dirs, files in os.walk(self.target_folder):
            for file in files:
                file_lower = file.lower()
                if any(p in file_lower for p in patterns):
                    full_path = os.path.join(root, file)
                    self.found_files.append(full_path)
                    item = QListWidgetItem(file)
                    self.ui.resultList.addItem(item)

    def show_preview(self, item):
        import glob
        import difflib
        from PyQt5.QtGui import QPixmap

        # Get the base name (without extension) of the selected item
        selected_name = os.path.splitext(item.text())[0].lower()

        # Search only in Products Image folder
        products_image_dir = os.path.join(self.server_path, "Images", "Products Image")
        image_files = []
        if os.path.exists(products_image_dir):
            image_files.extend(glob.glob(os.path.join(products_image_dir, "*.jpg")))
            image_files.extend(glob.glob(os.path.join(products_image_dir, "*.jpeg")))
            image_files.extend(glob.glob(os.path.join(products_image_dir, "*.png")))

        # Find the closest match by filename (without extension)
        image_basenames = [os.path.splitext(os.path.basename(f))[0].lower() for f in image_files]
        matches = difflib.get_close_matches(selected_name, image_basenames, n=1, cutoff=0.6)

        if matches:
            # Get the full path of the best match
            best_match_index = image_basenames.index(matches[0])
            best_image_path = image_files[best_match_index]
            pixmap = QPixmap(best_image_path)
            if not pixmap.isNull():
                scaled_pixmap = pixmap.scaled(
                    self.ui.previewLabel.width(),
                    self.ui.previewLabel.height(),
                    QtCore.Qt.KeepAspectRatio,
                    QtCore.Qt.SmoothTransformation
                )
                self.ui.previewLabel.setPixmap(scaled_pixmap)
                return

        # If no match or failed to load, clear the preview
        self.ui.previewLabel.clear()

    def download_file(self):
        from PyQt5.QtWidgets import QMessageBox
        
        selected_items = self.ui.resultList.selectedItems()
        if not selected_items:
            return
        file_name = selected_items[0].text()
        source_file = next((f for f in self.found_files if os.path.basename(f) == file_name), None)

        if source_file:
            target_folder = QFileDialog.getExistingDirectory(self, "Chọn thư mục để lưu")
            if target_folder:
                try:
                    shutil.copy(source_file, target_folder)
                except FileNotFoundError:
                    msg = QMessageBox(self)
                    msg.setIcon(QMessageBox.Warning)
                    msg.setWindowTitle("Lỗi")
                    msg.setText(self.tr("Không thể copy do file chưa đồng bộ. Mình sẽ dẫn bạn tới thư mục. "
                    "Hướng dẫn:\nChuột phải vào file, chọn Onedrive, chọn Copy Link, dán vào Chrome để mở và download"
                    "\nNếu không có Onedrive, Chuột phải vào file, chọn 'Always keep on this device'. Chờ đồng bộ xong rồi thử lại."))
                    msg.setStandardButtons(QMessageBox.Ok)
                    ret = msg.exec_()
                    if ret == QMessageBox.Ok:
                        folder = os.path.dirname(source_file)
                        os.startfile(folder)  # This opens the folder in Windows Explorer
                
    def print_image(self):
        selected_items = self.ui.resultList.selectedItems()
        if not selected_items:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Vui lòng chọn một file để in."))
            return
        file_name = selected_items[0].text()
        source_file = next((f for f in self.found_files if os.path.basename(f) == file_name), None)
        if not source_file:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy file."))
            return

        # Handle images
        if source_file.lower().endswith(('.png', '.jpg', '.jpeg')):
            from PyQt5.QtGui import QPixmap
            pixmap = QPixmap(source_file)
            if pixmap.isNull():
                QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không thể mở hình ảnh."))
                return
        # Handle PDFs
        elif source_file.lower().endswith('.pdf'):
            try:
                doc = fitz.open(source_file)
                page = doc.load_page(0)  # First page
                pix = page.get_pixmap(dpi=200)
                image_bytes = pix.tobytes("ppm")
                from PyQt5.QtGui import QImage, QPixmap
                image = QImage.fromData(image_bytes)
                pixmap = QPixmap.fromImage(image)
                if pixmap.isNull():
                    QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không thể mở file PDF."))
                    return
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), f"{self.tr('Không thể mở file PDF')}: {e}")
                return
        else:
            QtWidgets.QMessageBox.information(self, self.tr("Thông báo"), self.tr("Chỉ hỗ trợ in file hình ảnh (.jpg, .jpeg, .png) hoặc PDF."))
            return

        # Print the pixmap (image or PDF as image)
        printer = QPrinter()
        dialog = QPrintDialog(printer, self)
        if dialog.exec_() == QPrintDialog.Accepted:
            painter = QtGui.QPainter()
            if painter.begin(printer):
                page_rect = printer.pageRect()
                img_rect = pixmap.rect()
                scale = min(page_rect.width() / img_rect.width(), page_rect.height() / img_rect.height())
                x = (page_rect.width() - img_rect.width() * scale) / 2
                y = (page_rect.height() - img_rect.height() * scale) / 2
                painter.translate(x, y)
                painter.scale(scale, scale)
                painter.drawPixmap(0, 0, pixmap)
                painter.end()
        else:
            # Do nothing if print dialog is cancelled
            pass
    def open_colorway_ppt(self):
        import openpyxl
        from PyQt5.QtWidgets import QMessageBox
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        import difflib

        # Get the search keyword from searchLineEdit
        keyword = self.ui.searchLineEdit.text().strip()
        if not keyword:
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Vui lòng nhập từ khóa tìm kiếm."))
            return

        # Path to Excel file
        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        if not os.path.exists(excel_path):
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy file List Product QC.xlsx"))
            return

        # Open the workbook and sheet
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except PermissionError:
            QtWidgets.QMessageBox.warning(
                self,
                self.tr("Lỗi"),
                self.tr("Ai đó đang mở file excel List Product QC.xlsx. Hãy tắt trước khi thao tác")
            )
            return

        if "RenderColorway" not in wb.sheetnames:
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy sheet 'RenderColorway'"))
            return
        ws = wb["RenderColorway"]

        # Find all colorway values and product names
        header = [cell.value for cell in ws[1]]
        try:
            col_products = header.index("Products")
            col_colorway = header.index("Color way")
        except ValueError:
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy cột 'Products' hoặc 'Color way'"))
            return

        product_list = []
        colorway_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            product = str(row[col_products]).strip() if row[col_products] else ""
            colorway = str(row[col_colorway]).strip() if row[col_colorway] else ""
            if product and colorway:
                product_list.append(product)
                colorway_list.append(colorway)

        # Find nearest matches in Products column using difflib
        # Lowercase all product names for case-insensitive matching
        product_list_lower = [p.lower() for p in product_list]
        matches_lower = difflib.get_close_matches(keyword.lower(), product_list_lower, n=10, cutoff=0.3)
        # Map back to original case
        matches = [product_list[product_list_lower.index(m)] for m in matches_lower]
        
        if not matches:
            QMessageBox.information(self, self.tr("Kết quả"), self.tr("Không tìm thấy sản phẩm phù hợp."))
            return

        # Show dialog with product matches
        dialog = ColorwayDialog(matches, self)
        if dialog.exec_() == QDialog.Accepted:
            selected_product = dialog.get_selected_colorway()
            if not selected_product:
                QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Vui lòng chọn một sản phẩm."))
                return

            # Find the corresponding Color way for the selected product
            try:
                idx = product_list.index(selected_product)
                selected_colorway = colorway_list[idx]
            except ValueError:
                QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy Color way cho sản phẩm đã chọn."))
                return

            # Build the path to the PowerPoint file
            ppt_folder = os.path.join(
                self.company_folder,
                "TLC DRAWINGS",
                "STANDARD DRAWNGS  SEND TO SUPPLIER",
                "RENDERING COLOR WAY"
            )
            ppt_path = os.path.join(ppt_folder, f"{selected_colorway}.pptx")

            if not os.path.exists(ppt_path):
                QMessageBox.warning(self, self.tr("Lỗi"), self.tr(f"Không tìm thấy file PowerPoint: {ppt_path}"))
                return

            # Open the PowerPoint file
            try:
                os.startfile(ppt_path)  # Windows only
            except Exception as e:
                QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể mở file PowerPoint')}: {e}")
    
    def open_selected_file(self):
        from PyQt5.QtWidgets import QMessageBox

        selected_items = self.ui.resultList.selectedItems()
        if not selected_items:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Vui lòng chọn một file để mở."))
            return
        file_name = selected_items[0].text()
        source_file = next((f for f in self.found_files if os.path.basename(f) == file_name), None)

        if source_file and os.path.exists(source_file):
            try:
                os.startfile(source_file)  # Windows only
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể mở file')}: {e}")
                msg = QMessageBox(self)
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle("Lỗi")
                msg.setText(self.tr("Không thể copy do file chưa đồng bộ. Mình sẽ dẫn bạn tới thư mục."
                "Hướng dẫn:\nChuột phải vào file, chọn Onedrive, chọn Copy Link, dán vào Chrome để mở và download"
                "\nNếu không có Onedrive, Chuột phải vào file, chọn 'Always keep on this device'. Chờ đồng bộ xong rồi thử lại."))
                msg.setStandardButtons(QMessageBox.Ok)
                ret = msg.exec_()
                if ret == QMessageBox.Ok:
                    folder = os.path.dirname(source_file)
                    os.startfile(folder)
        else:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy file."))
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Lỗi")
            msg.setText(self.tr("Không thể copy do file chưa đồng bộ. Mình sẽ dẫn bạn tới thư mục. "
                    "Hướng dẫn:\nChuột phải vào file, chọn Onedrive, chọn Copy Link, dán vào Chrome để mở và download"
                    "\nNếu không có Onedrive, Chuột phải vào file, chọn 'Always keep on this device'. Chờ đồng bộ xong rồi thử lại."))
            msg.setStandardButtons(QMessageBox.Ok)
            ret = msg.exec_()
            if ret == QMessageBox.Ok:
                folder = os.path.dirname(source_file)
                os.startfile(folder)

class timsp(QtWidgets.QMainWindow):
    def __init__(self, server_path):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.server_path = server_path
        self.search_mode = "Final"  # Default mode
        # QCompleter for product names
        import openpyxl
        from PyQt5.QtWidgets import QCompleter
        
        # Connect the buttons
        self.ui.TimSPButton.clicked.connect(self.handle_timsp_button)
        self.ui.TCKiemtraButton.clicked.connect(self.show_hangmuc_results) 
        self.ui.lineEdit.returnPressed.connect(self.handle_timsp_button)
        
        matrix_sheet = "Matrix"
        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        try:
            with open(excel_path, 'rb') as f:
                pass
        except PermissionError:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Ai đó đang mở file excel List Product QC.xlsx. Hãy tắt trước khi thao tác"))
            return
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể mở file Excel')}: {e}")
            return
        
        product_names = set()
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            if matrix_sheet in wb.sheetnames:
                ws = wb[matrix_sheet]
                for row in ws.iter_rows(min_row=5, min_col=2, max_col=2):
                    cell = row[0]
                    if cell.value:
                        product_names.add(str(cell.value).strip())

        completer = QCompleter(sorted(product_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit.setCompleter(completer)
    
    def handle_timsp_button(self):
        self.show_search_mode_dialog()
        self.search_product()

    def show_search_mode_dialog(self):
        dialog = QtWidgets.QDialog(self)
        dialog.setFixedWidth(250)
        dialog.setWindowTitle(self.tr("Chọn loại"))
        layout = QtWidgets.QVBoxLayout(dialog)
        btn_carcass = QtWidgets.QPushButton(self.tr("Carcass"))
        btn_final = QtWidgets.QPushButton(self.tr("Final"))
        layout.addWidget(btn_carcass)
        layout.addWidget(btn_final)
        
        btn_final.clicked.connect(lambda: (setattr(self, "search_mode", "Final"), dialog.accept()))
        btn_carcass.clicked.connect(lambda: (setattr(self, "search_mode", "Carcass"), dialog.accept()))
    
        dialog.exec_()
    
    def search_product(self):
        import openpyxl
        import os
        from PyQt5.QtGui import QStandardItemModel, QStandardItem, QFont

        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        if not os.path.exists(excel_path):
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr("Could not find List Product QC.xlsx"))
            return
        try:
            with open(excel_path, 'rb') as f:
                pass
        except PermissionError:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Ai đó đang mở file excel List Product QC.xlsx. Hãy tắt trước khi thao tác"))
            return
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể mở file Excel')}: {e}")
            return
        
        keyword = self.ui.lineEdit.text().strip().lower()
        if not keyword:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Vui lòng nhập từ khóa tìm kiếm."))
            return
    
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if self.search_mode == "Carcass":
            sheet_name = "MatrixCarcass"
        else:
            sheet_name = "Matrix"

        if sheet_name not in wb.sheetnames:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr(f"Không tìm thấy sheet '{sheet_name}' trong file."))
            return
        ws = wb[sheet_name]
    
        # Get header and columns to show
        if self.search_mode == "Carcass":
            # Columns B to G (2 to 7)
            header = [ws.cell(row=4, column=col).value for col in range(2, 8)]
            min_col, max_col = 2, 7
        else:
            # Columns B to M (2 to 13)
            header = [ws.cell(row=4, column=col).value for col in range(2, 14)]
            min_col, max_col = 2, 13
    
        # Search for results in the first column of the range, from row 5
        results = []
        for row in ws.iter_rows(min_row=5, min_col=min_col, max_col=min_col):
            cell = row[0]
            value = str(cell.value).strip().lower() if cell.value else ""
            if keyword in value:
                # Get the whole row in the range
                row_data = [ws.cell(row=cell.row, column=col).value for col in range(min_col, max_col + 1)]
                results.append(row_data)
    
        # Set up the model for QTableView
        model = QStandardItemModel()
        model.setColumnCount(len(header))
        model.setHorizontalHeaderLabels([str(h) if h is not None else "" for h in header])
    
        font = QFont()
        font.setBold(True)
        font.setPointSize(12)
        for col in range(model.columnCount()):
            model.setHeaderData(col, QtCore.Qt.Horizontal, font, QtCore.Qt.FontRole)
    
        for row_data in results:
            items = []
            for value in row_data:
                item = QStandardItem(str(value) if value is not None else "")
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                items.append(item)
            model.appendRow(items)
    
        self.ui.tableView.setModel(model)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
    
        if not results:
            QtWidgets.QMessageBox.information(self, self.tr("Kết quả"), self.tr("Không tìm thấy sản phẩm phù hợp."))
    
        self.last_matrix_results = results
        self.last_matrix_header = header
        search_mode = self.search_mode  # or however you store it


    def show_hangmuc_results(self):
        import openpyxl
        from PyQt5.QtGui import QStandardItemModel, QStandardItem, QFont

        # Get last search results and header
        matrix_header = getattr(self, "last_matrix_header", None)
        if not matrix_header:
            QtWidgets.QMessageBox.information(self, self.tr("Thông báo"), self.tr("Chưa có dữ liệu kiểm tra. Vui lòng tìm kiếm trước."))
            return

        # Get selected row in tableView
        selection_model = self.ui.tableView.selectionModel()
        if not selection_model.hasSelection():
            QtWidgets.QMessageBox.information(self, self.tr("Thông báo"), self.tr("Vui lòng chọn một dòng kết quả để kiểm tra."))
            return
        selected_indexes = selection_model.selectedRows()
        if not selected_indexes:
            QtWidgets.QMessageBox.information(self, self.tr("Thông báo"), self.tr("Vui lòng chọn một dòng kết quả để kiểm tra."))
            return
        selected_row = selected_indexes[0].row()

        # Get the data of the selected row
        model = self.ui.tableView.model()
        row_data = []
        for col in range(model.columnCount()):
            item = model.item(selected_row, col)
            row_data.append(item.text() if item else "")

        # Open Excel and get the correct Hangmuc sheet
        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if getattr(self, "search_mode", "Final") == "Carcass":
            sheet_name = "HangmucCarcass"
        else:
            sheet_name = "Hangmuc"

        if sheet_name not in wb.sheetnames:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr(f"Không tìm thấy sheet '{sheet_name}' trong file."))
            return
        ws_hangmuc = wb[sheet_name]
        # Prepare results: [(header, [B, C, D, E, F])]
        hangmuc_results = []
        for col_idx, cell_value in enumerate(row_data):
            if str(cell_value).strip().lower() == "o":
                header_value = matrix_header[col_idx]
                # Collect ALL rows in Hangmuc where column A matches header_value
                for hangmuc_row in ws_hangmuc.iter_rows(min_row=1, min_col=1, max_col=6):
                    if str(hangmuc_row[0].value).strip() == str(header_value).strip():
                        hangmuc_row_data = [cell.value for cell in hangmuc_row[1:6]]  # B, C, D, E, F
                        hangmuc_results.append([header_value] + hangmuc_row_data)

        # Show in dialog
        dialog = TCKiemtraDialog(self, searched_value=self.ui.lineEdit.text(),search_mode = self.search_mode)

        model = QStandardItemModel()
        model.setColumnCount(6)  # Columns B to F (1 to 5 in Excel)
        # Get header from row 2, columns A to F (1 to 6)
        hangmuc_header = [ws_hangmuc.cell(row=2, column=col).value for col in range(1, 7)]
        model.setHorizontalHeaderLabels([str(h) if h is not None else "" for h in hangmuc_header])

        font = QFont()
        font.setBold(True)
        font.setPointSize(12)
        for col in range(model.columnCount()):
            model.setHeaderData(col, QtCore.Qt.Horizontal, font, QtCore.Qt.FontRole)

        for row in hangmuc_results:
            items = []
            for value in row:
                item = QStandardItem(str(value) if value is not None else "")
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                items.append(item)
            model.appendRow(items)

        dialog.ui.tableView.setModel(model)
        dialog.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

        # Show image when a row is selected
        def show_image_for_row(row):
            if row < 0:
                dialog.ui.imageLabel.clear()
                return
            value_in_D = model.item(row, 5).text() if model.item(row, 5) else ""
            if value_in_D:
                image_path = os.path.join(self.server_path, "Images", "Defect Images", f"{value_in_D}.jpg")
                if os.path.exists(image_path):
                    pixmap = QtGui.QPixmap(image_path)
                    if not pixmap.isNull():
                        scaled = pixmap.scaled(400, 400, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
                        dialog.ui.imageLabel.setPixmap(scaled)
                        return
            dialog.ui.imageLabel.clear()

        dialog.ui.tableView.selectionModel().currentRowChanged.connect(lambda current, previous: show_image_for_row(current.row()))
        if model.rowCount() > 0:
            dialog.ui.tableView.selectRow(0)
        dialog.exec_()
        

class TCKiemtraDialog(QDialog):
    def __init__(self, parent=None, searched_value="",search_mode=None):
        super().__init__(parent)
        self.ui = Ui_TCKiemtraDialog()
        self.ui.setupUi(self)
        self.ui.exportexcelButton.clicked.connect(self.export_to_excel)
        
        # Set lineEdit_2 with the searched value from timsp
        self.ui.lineEdit_2.setText(searched_value)
        self.search_mode = search_mode

    def export_to_excel(self):
        import openpyxl
        from PyQt5.QtWidgets import QMessageBox

        # Path to the template Excel file
        if self.search_mode == "Carcass":
            excel_path = os.path.join(self.parent().server_path, "QC Check list - Format carcass - vi.xlsx")
            sheet_name = "Format_car_vi"
        else:
            excel_path = os.path.join(self.parent().server_path, "QC Check list - Format.xlsx")
            sheet_name = "Format_fin_vi"
        
        if not os.path.exists(excel_path):
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy file QC Check list - Format.xlsx"))
            return

        # Open the workbook and select the correct sheet
        wb = openpyxl.load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr(f"Không tìm thấy sheet '{sheet_name}'"))
            return
        ws = wb[sheet_name]

        # Get data from tableView
        model = self.ui.tableView.model()
        if model is None or model.rowCount() == 0:
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không có dữ liệu để xuất."))
            return

        # Write data to the Excel sheet
        for row in range(model.rowCount()):
            # index 0 -> column B, index 1 -> column A, index 2 -> column D, index 3 -> column G, index 4 -> column L
            ws.cell(row=9+row, column=2).value = model.item(row, 0).text() if model.item(row, 0) else ""
            ws.cell(row=9+row, column=1).value = model.item(row, 1).text() if model.item(row, 1) else ""
            ws.cell(row=9+row, column=4).value = model.item(row, 2).text() if model.item(row, 2) else ""
            ws.cell(row=9+row, column=7).value = model.item(row, 3).text() if model.item(row, 3) else ""
            ws.cell(row=9+row, column=12).value = model.item(row, 4).text() if model.item(row, 4) else ""
        
        # Assign lineEdit and lineEdit_2 values to specific cells
        ws["C4"].value = self.ui.lineEdit.text().upper()
        ws["L4"].value = self.ui.lineEdit_2.text().upper()
        ws["G4"].value = self.ui.lineEdit_3.text().upper()

        # Calculate the last row you wrote to
        last_row = 9 + model.rowCount() - 1  # Data starts at row 9

        # Set the print area: columns A to L (1 to 12), rows 9 to last_row
        ws.print_area = f"A1:P{last_row}"
        
        # Save as a new file (ask user where to save)
        from PyQt5.QtWidgets import QFileDialog
        # Get values from lineEdit and lineEdit_2 for the filename
        name1 = self.ui.lineEdit.text().upper().strip()
        name2 = self.ui.lineEdit_2.text().upper().strip()
        name3 = self.ui.lineEdit_3.text().upper().strip()
        # Sanitize filename (remove or replace invalid characters)
        import re
        def sanitize(s):
            return re.sub(r'[\\/*?:"<>|]', "_", s)
        filename = f"{sanitize(name1)}_{sanitize(name2)}_{sanitize(name3)}.xlsx" if name1 or name2 or name3 else "QC Check list - export.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(self, self.tr("Lưu file Excel"), filename, "Excel Files (*.xlsx)")

        def sanitize1(s):
            # Excel sheet names can't have: : \ / ? * [ ]
            return re.sub(r'[:\\/*?\[\]]', "_", s)[:31]  # Excel sheet name max length is 31

        ws.oddHeader.center.text = f"{sanitize1(name1)}_{sanitize1(name2)}_{sanitize1(name3)}" if name1 or name2 or name3 else "Sheet1"
        # Rename the worksheet
        ws.title = sheet_name

        if save_path:
            try:
                wb.save(save_path)
                QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã xuất dữ liệu ra file Excel thành công!"))
            except Exception as e:
                QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể lưu file Excel')}: {e}")

class ColorwayDialog(QDialog):
    def __init__(self, results, parent=None):
        super().__init__(parent)
        self.ui = Ui_ColorwayDialog()
        self.ui.setupUi(self)

        # Fill the listView with results
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        self.model = QStandardItemModel()
        for product in results:
            item = QStandardItem(product)
            self.model.appendRow(item)
        self.ui.listView.setModel(self.model)

        # Connect OK and Cancel buttons
        self.ui.okButton.clicked.connect(self.accept)
        self.ui.cancelButton.clicked.connect(self.reject)

    def get_selected_colorway(self):
        indexes = self.ui.listView.selectedIndexes()
        if indexes:
            return self.model.item(indexes[0].row()).text()
        return None
    
class CollectionDialog(QDialog):
    COLLECTION_MAP = {
        "MILESREDDButton": "MILES REDD",
        "SUZANNESHARPButton": "SUZANNE SHARP",
        "VEEREGRENNEYButton": "VEERE GRENNEY",
        "PETERMIKICButton": "PETER MIKIC",
        "RITAKONIGButton": "RITA KONIG",
        "KRBButton": "KRB",
        "JANECHURCHILLButton": "JANE CHURCHILL",
        "PENTREATHHALLButton": "PENTREATH & HALL",
        "STEVENGAMBRELButton": "STEVEN GAMBREL",
        "LUKEEDWARDHALLButton": "LUKE EDWARD HALL",
        "JOHNDERIANButton": "JOHN DERIAN",
        "HOWELONDONButton": "HOWE LONDON",
        "THELACQUERCOMPANYButton": "THE LACQUER COMPANY",
        "SCHUMACHERButton": "SCHUMACHER",
        "CHRISTOPHERSPITZMILLERButton": "CHRISTOPHER SPITZMILLER",
        "SALVESENGRAHAMButton": "SALVESEN GRAHAM",
        "JEFFREYBILHUBERButton": "JEFFREY BILHUBER",
        "CAMPBELLREYButton": "CAMPBELL-REY",
    }

    def __init__(self, server_path, parent=None):
        super().__init__(parent)
        self.ui = Ui_CollectionDialog()
        self.ui.setupUi(self)

        self.server_path = server_path

        # Load Excel data once
        import pandas as pd
        import os

        self.df = None
        try:
            excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
            self.df = pd.read_excel(excel_path, sheet_name="DATA")
            with open(excel_path, 'rb') as f:
                pass
        except PermissionError:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Ai đó đang mở file excel List Product QC.xlsx. Hãy tắt trước khi thao tác"))
            return
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể mở file Excel')}: {e}")
            return
        
        # Connect all buttons to the same handler
        for btn_name, collection_value in self.COLLECTION_MAP.items():
            btn = getattr(self.ui, btn_name, None)
            if btn:
                btn.clicked.connect(lambda checked, col=collection_value: self.show_collection_products(col))

    def show_collection_products(self, collection_value):
        if self.df is None:
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr("No data loaded."))
            return
        
        # Filter rows matching the collection
        filtered = self.df[self.df["COLLECTION"].astype(str).str.strip().str.upper() == collection_value.upper()]

        # Drop rows where PRODUCT NAME or ITEM NO. is missing
        filtered = filtered.dropna(subset=["PRODUCT NAME", "ITEM NO.","COLOR"])

        # Get unique pairs of (PRODUCT NAME, ITEM NO.)
        products = filtered[["PRODUCT NAME", "ITEM NO.","COLOR"]].drop_duplicates().sort_values(by=["PRODUCT NAME", "ITEM NO.", "COLOR"])

        # Show in tableView
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        model = QStandardItemModel()
        model.setHorizontalHeaderLabels(["PRODUCT NAME", "ITEM NO.", "COLOR"])
        for _, row in products.iterrows():
            item_name = QStandardItem(str(row["PRODUCT NAME"]))
            item_no = QStandardItem(str(row["ITEM NO."]))
            item_color = QStandardItem(str(row["COLOR"]))
            item_name.setTextAlignment(QtCore.Qt.AlignCenter)
            item_no.setTextAlignment(QtCore.Qt.AlignCenter)
            item_color.setTextAlignment(QtCore.Qt.AlignCenter)
            model.appendRow([item_name, item_no, item_color])
        self.ui.tableView.setModel(model)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

        # Set minimum width for "PRODUCT NAME" (column 0) and "COLOR" (column 2)
        self.ui.tableView.setColumnWidth(0, max(240, self.ui.tableView.columnWidth(0)))
        self.ui.tableView.setColumnWidth(2, max(270, self.ui.tableView.columnWidth(2)))
        self.ui.tableView.horizontalHeader().setMinimumSectionSize(130)  # Optional: set a general minimum

class defectlist(QtWidgets.QWidget):
    def __init__(self, server_path, parent=None):
        super().__init__(parent)
        from PyQt5 import QtWidgets, QtCore
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        import openpyxl
        import os
        self.ui = Ui_DefectListWidget()
        self.ui.setupUi(self)
        self.server_path = server_path

        excel_path = os.path.join(server_path, "Defects list.xlsx")
        if not os.path.exists(excel_path):
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không tìm thấy file: {excel_path}"))
            return

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except PermissionError:
            QtWidgets.QMessageBox.warning(
                self,
                self.tr("Lỗi"),
                self.tr("Ai đó đang mở file excel List Product QC.xlsx. Hãy tắt trước khi thao tác")
            )
            return
        ws = wb["Defect list"]

        # Read columns A to J (1 to 10)
        data = []
        for row in ws.iter_rows(min_row=1, max_col=10, values_only=True):
            data.append([str(cell) if cell is not None else "" for cell in row])

        model = QStandardItemModel()
        model.setColumnCount(10)
        if data:
            model.setHorizontalHeaderLabels([str(h) for h in data[0]])
            for row in data[1:]:
                items = [QStandardItem(cell) for cell in row]
                for item in items:
                    item.setTextAlignment(QtCore.Qt.AlignCenter)
                model.appendRow(items)

        self.ui.tableView.setModel(model)
        self.ui.tableView.setAlternatingRowColors(True)
        self.ui.tableView.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
        self.ui.tableView.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.ui.tableView.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.ui.tableView.setShowGrid(True)
        self.ui.tableView.verticalHeader().setVisible(False)
        self.ui.defectaddopenButton.clicked.connect(self.open_defectinput_widget)

    def open_defectinput_widget(self):
        excel_path = os.path.join(self.server_path, "Defects list.xlsx")
        try:
            with open(excel_path, 'rb') as f:
                pass
            self.defect_input_widget = defectinput(self.server_path)
            self.defect_input_widget.show()
            return
        except PermissionError:
            QtWidgets.QMessageBox.warning(
                self,
                self.tr("Lỗi"),
                self.tr("Ai đó đang mở file excel Defects list.xlsx. Hãy tắt trước khi thao tác")
            )
            return
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, 
                self.tr("Lỗi"), 
                self.tr(f"Đã xảy ra lỗi không mong muốn khi mở widget hoặc file: {e}")
            )
            return


class defectinput(QtWidgets.QWidget):
    def __init__(self, server_path, parent=None):
        super().__init__(parent)
        self.ui = Ui_DefectInputWidget()
        self.ui.setupUi(self)
        self.server_path = server_path

        self.ui.adddefectButton.clicked.connect(self.adddefectlist)

    def adddefectlist(self):
        import openpyxl
        import os

        # Path to your Excel file (adjust as needed)
        excel_path = os.path.join(self.server_path, "Defects list.xlsx")

        try:
            with open(excel_path, 'rb') as f:
                pass
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["Defect list"]
            next_row = ws.max_row + 1

            ws.cell(row=next_row, column=1, value=self.ui.comboBox.currentText())     # A (QComboBox)
            ws.cell(row=next_row, column=2, value=self.ui.lineEdit_2.text())          # B
            ws.cell(row=next_row, column=7, value=self.ui.lineEdit_3.text())          # G
            ws.cell(row=next_row, column=3, value=self.ui.lineEdit_4.text())          # C
            ws.cell(row=next_row, column=4, value=self.ui.lineEdit_5.text())          # D
            ws.cell(row=next_row, column=5, value=self.ui.lineEdit_6.text())          # E
            ws.cell(row=next_row, column=9, value=self.ui.comboBox_2.currentText())   # I
            ws.cell(row=next_row, column=10, value=self.ui.comboBox_3.currentText())  # J

            # QDateEdit for column F (6), format "dd-mm-yyyy"
            date_str = self.ui.dateEdit.date().toString("dd-MM-yy")
            ws.cell(row=next_row, column=6, value=date_str)  # F

            wb.save(excel_path)
            wb.close()
            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã thêm lỗi vào danh sách!"))
            for i in range(2, 7):
                getattr(self.ui, f"lineEdit_{i}").setText("")
        except PermissionError:
            QtWidgets.QMessageBox.warning(
                self,
                self.tr("Lỗi"),
                self.tr("Ai đó đang mở file excel Defects list.xlsx. Hãy tắt trước khi thao tác")
            )
            return
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr(f"Không thể ghi file Excel:\n{e}"))

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QIcon(":/images/tlclogo4.png"))
    translator = QTranslator()
    lang_code = "vi"  # or "en", or use a config/setting
    translator.load(f"app_{lang_code}.qm")
    app.installTranslator(translator)
        
    window = ColorSearchApp()
    window.setWindowTitle(window.tr("THE LACQUER COMPANY APP"))
    window.setWindowIcon(QIcon(":/images/tlclogo4.png"))
    window.show()
    sys.exit(app.exec_())