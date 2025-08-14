import sys
import os
import urllib.parse
import urllib.request
import pandas as pd
from PyQt5 import QtWidgets
from PyQt5 import QtCore
from form import Ui_Form  # form.py sinh ra từ form.ui
from color_form import Ui_Form as Ui_ColorDialog  # Adjust class name if needed
from effect_form import Ui_Form as Ui_EffectDialog  # Adjust class name if needed
from metal_form import Ui_Form as Ui_MetalDialog  # Adjust class name if needed
from wood_form import Ui_Form as Ui_WoodDialog  # Adjust class name and filename if needed
import shutil  # Add at the top of your file if not already imported
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QDialogButtonBox, QLabel, QRadioButton, QButtonGroup, QListWidgetItem, QFileDialog
from bvstd_window import Ui_Dialog 
from timsp import Ui_MainWindow
from PyQt5.QtPrintSupport import QPrintDialog, QPrinter
from PyQt5 import QtGui
import fitz  # PyMuPDF


def find_company_folder():
    # List all possible company folder paths
    possible_paths = [
        r"C:\Users\Admins\The Lacquer Company\Company Files - Tài liệu",
        r"C:\Users\Admins\The Lacquer Company\Company Files - Documents"
    ]
    for path in possible_paths:
        if os.path.exists(path):
            return path
    raise RuntimeError("Không tìm thấy thư mục công ty trên máy tính này.")

def all_words_in_text(words, text):
    return all(word in text for word in words)

class ColorSearchApp(QtWidgets.QWidget):    
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.company_folder = find_company_folder()
        self.server_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "QC FOLDER",
            "MASTER COLOR LIST QC"
        )
        serverexcel_path = os.path.join(self.server_path,"TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")

        local_path = 'TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx'

        try:
            self.df_main = pd.read_excel(serverexcel_path, sheet_name='Lacquer FIN')
            self.df_custom = pd.read_excel(serverexcel_path, sheet_name='Custom color')
            self.df_metal = pd.read_excel(serverexcel_path, sheet_name='Metal FIN')
            self.df_wood = pd.read_excel(serverexcel_path, sheet_name='Wood FIN')          
            self.df_effect = pd.read_excel(serverexcel_path, sheet_name='Effect Color Swatch Statistics')
        except Exception:
            self.df_main = pd.read_excel(local_path, sheet_name='Lacquer FIN')
            self.df_custom = pd.read_excel(local_path, sheet_name='Custom color')
            self.df_metal = pd.read_excel(local_path, sheet_name='Metal FIN')
            self.df_wood = pd.read_excel(local_path, sheet_name='Wood FIN')
            self.df_effect = pd.read_excel(local_path, sheet_name='Effect Color Swatch Statistics')

        # Strip spaces from column names for robustness
        self.df_main.columns = self.df_main.columns.str.strip()
        self.df_custom.columns = self.df_custom.columns.str.strip()
        self.df_metal.columns = self.df_metal.columns.str.strip()
        self.df_wood.columns = self.df_wood.columns.str.strip()
        self.df_effect.columns = self.df_effect.columns.str.strip()

        self.ui.pushButton.clicked.connect(self.search_color)
        self.ui.HieuUngButton.clicked.connect(self.search_effect_color)  # Connect the new button
        self.ui.ThemButton.clicked.connect(self.show_option_dialog)
        self.ui.BVSTDButton.clicked.connect(self.open_bvstd_window)
        self.ui.TCQCButton.clicked.connect(self.open_timsp_window)

        self.ui.lineEdit.returnPressed.connect(self.search_color)  # Allow pressing Enter to search
    
        # After loading all DataFrames in __init__:
        self.all_product_names = set()
        for df in [self.df_main, self.df_custom, self.df_metal, self.df_wood]:
            if "Name" in df.columns:
                self.all_product_names.update(df["Name"].dropna().astype(str).tolist())

        from PyQt5.QtWidgets import QCompleter

        completer = QCompleter(sorted(self.all_product_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit.setCompleter(completer)

    def open_bvstd_window(self):
        self.bvstd_window = BVSTDWindow(self.company_folder, self.server_path)
        self.bvstd_window.exec_()

    def open_timsp_window(self):
        self.timsp_window = timsp(self.server_path)
        self.timsp_window.show()

    def search_color(self):
        keyword = self.ui.lineEdit.text().strip().lower()

        # Search in 'Lacquer FIN'
        words = keyword.split()
        matched = self.df_main[self.df_main['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
        
        # If not found, search in 'Custom color'
        if matched.empty:
            #words = keyword.split()
            matched = self.df_custom[self.df_custom['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
            
        if matched.empty:
            #words = keyword.split()
            matched = self.df_metal[self.df_metal['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
        
        if matched.empty:
            #words = keyword.split()
            matched = self.df_wood[self.df_wood['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
            if matched.empty:
                self.ui.textEdit.setText("🔍 Không tìm thấy kết quả.")
                return

        # Add the first sentence with bold and underline
        
        html=""
        for _, row in matched.iterrows():
            name = str(row.get("Name", "")).strip()
            for col in matched.columns:
                cell = row.get(col, "")
                # If cell is a Series (shouldn't be, but just in case), get the first value
                if isinstance(cell, pd.Series):
                    cell = cell.iat[0] if not cell.empty else ""
                try:
                    is_na = pd.isna(cell)
                except Exception:
                    is_na = False
                value = "" if is_na else str(cell)
                if col.strip().lower() in ["ref image"]:
                    image_path = os.path.join(self.server_path, "Images", f"{name}.jpg")
                    if os.path.exists(image_path):
                        file_uri = urllib.parse.urljoin('file:', urllib.request.pathname2url(image_path))
                        html += f'<p><b>{col}:</b><br><img src="{file_uri}" width="200"></p>'
                    else:
                        html += f"<p><b>{col}:</b> (Không tìm thấy ảnh)</p>"
                else:
                    html += f"<p><b>{col}:</b> {value}</p>"
            html += "<hr>"

        self.ui.textEdit.setHtml(html)

    def search_effect_color(self):
        keyword = self.ui.lineEdit.text().strip().lower()
        # Adjust the column name below if needed
        if 'Color Name' in self.df_effect.columns:
            words = keyword.split()
            matched = self.df_effect[self.df_effect['Color Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
        else:
            # Print columns for debugging if needed
            print(self.df_effect.columns.tolist())
            self.ui.textEdit.setText("Không tìm thấy cột 'Color Name' trong sheet hiệu ứng.")
            return

        if matched.empty:
            self.ui.textEdit.setText("🔍 Không tìm thấy kết quả trong hiệu ứng.")
            return

        html = '<p><b><u>EFFECT COLOR</u></b></p>'
        for _, row in matched.iterrows():
            name=str(row.get("Color Name","")).strip()
            for col in matched.columns:
                cell = row.get(col, "")
                value = str(cell) if not pd.isna(cell) else ""
                if col.strip().lower() in ["ref image"]:
                    image_path=os.path.join(self.server_path,"Images",f"{name}.jpg")
                    if os.path.exists(image_path):
                        file_uri=urllib.parse.urljoin('file:',urllib.request.pathname2url(image_path))
                        html += f'<p><b>{col}:</b><br><img src="{file_uri}" width="200"></p>'
                    else:
                        html += f"<p><b>{col}:</b> (Không tìm thấy ảnh)</p>"
                else:
                    html += f"<p><b>{col}:</b> {value}</p>"
            html += "<hr>"

        self.ui.textEdit.setHtml(html)

    def show_option_dialog(self):
        dialog = OptionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            selected = dialog.selected_option()
            if selected == "Color":
                color_dialog = ColorFormDialog(self)
                color_dialog.exec_()
            elif selected == "Effect":
                effect_dialog = EffectFormDialog(self)
                effect_dialog.exec_()
            elif selected == "Metal":
                metal_dialog = MetalFormDialog(self)
                metal_dialog.exec_()
            elif selected == "Wood":
                wood_dialog = WoodFormDialog(self)
                wood_dialog.exec_()
            else:
                QtWidgets.QMessageBox.information(self, "Bạn đã chọn", f"Bạn đã chọn: {selected}")


class OptionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Chọn loại")
        layout = QVBoxLayout(self)

        label = QLabel("Chọn một loại:")
        layout.addWidget(label)

        self.button_group = QButtonGroup(self)
        self.radio_color = QRadioButton("Color")
        self.radio_metal = QRadioButton("Metal")
        self.radio_wood = QRadioButton("Wood")
        self.radio_effect = QRadioButton("Effect")
        self.button_group.addButton(self.radio_color)
        self.button_group.addButton(self.radio_metal)
        self.button_group.addButton(self.radio_wood)
        self.button_group.addButton(self.radio_effect)
        layout.addWidget(self.radio_color)
        layout.addWidget(self.radio_metal)
        layout.addWidget(self.radio_wood)
        layout.addWidget(self.radio_effect)

        self.radio_color.setChecked(True)  # Default selection

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_option(self):
        if self.radio_color.isChecked():
            return "Color"
        elif self.radio_metal.isChecked():
            return "Metal"
        elif self.radio_wood.isChecked():
            return "Wood"
        elif self.radio_effect.isChecked():
            return "Effect"
        return None

class ColorFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_ColorDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        self.them_mau_button = self.ui.ThemMauButton
        self.them_mau_button.clicked.connect(self.add_new_color)
        self.up_anh_button = self.ui.UpAnhButton
        self.up_anh_button.clicked.connect(self.upload_image)
        self.uploaded_image_path = None  # <-- Add this line

    def add_new_color(self):
        # Map lineEdits to columns
        mapping = [
            ("lineEdit_1", "Collection"),
            ("lineEdit_2", "Ref-Tone Code"),
            ("lineEdit_3", "Name"),
            ("lineEdit_4", "Reference"),
            ("lineEdit_5", "Status"),
            ("lineEdit_6", "Generation"),
            ("lineEdit_7", "Process"),
            ("lineEdit_8", "Request day"),
            ("lineEdit_9", "Qty"),
            ("lineEdit_10", "Approved by "),
            ("lineEdit_11", "Approved day"),
            ("lineEdit_12", "Sup - incharge"),
            ("lineEdit_13", "Master"),
            ("lineEdit_14", "Notes"),
            ("lineEdit_15", "Applied"),
        ]

        # Read values from lineEdits
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        # Load Excel file
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Lacquer FIN")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", f"Không thể đọc file Excel: {e}")
            return

        # Append new row
        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        # Save back to Excel
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Lacquer FIN", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from "Name" as the filename
                image_name = row_data.get("Name", "new_image").strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, "Thành công", "Đã thêm màu mới vào danh sách.")
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", f"Không thể ghi file Excel: {e}")

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Chọn ảnh JPG",
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path  # <-- Save the path

class EffectFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_EffectDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        self.uploaded_image_path = None

        self.ui.UpAnhHUButton.clicked.connect(self.upload_image)
        self.ui.ThemHUButton.clicked.connect(self.add_new_effect)

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Chọn ảnh JPG",
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path

    def add_new_effect(self):
        mapping = [
            ("lineEdit_1", "Color Name"),
            ("lineEdit_2", "Qty"),
            ("lineEdit_3", "Approval date"),
            ("lineEdit_4", "Note"),
        ]
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Effect Color Swatch Statistics")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", f"Không thể đọc file Excel: {e}")
            return

        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Effect Color Swatch Statistics", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from lineEdit_1 as the filename
                image_name = self.ui.lineEdit_1.text().strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, "Thành công", "Đã thêm hiệu ứng mới vào danh sách.")
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", f"Không thể ghi file Excel: {e}")

class MetalFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_MetalDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        self.uploaded_image_path = None

        self.ui.UpAnhMeButton.clicked.connect(self.upload_image)
        self.ui.ThemMetalButton.clicked.connect(self.add_new_metal)

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Chọn ảnh JPG",
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path

    def add_new_metal(self):
        mapping = [
            ("lineEdit_1", "Code"),
            ("lineEdit_2", "Name"),
            ("lineEdit_3", "Status"),
            ("lineEdit_4", "Description"),
            ("lineEdit_5", "Generation"),
            ("lineEdit_6", "Process"),
            ("lineEdit_7", "Request day"),
            ("lineEdit_8", "Qty"),
            ("lineEdit_9", "Approved by"),
            ("lineEdit_10", "Approved day"),
            ("lineEdit_11", "Reject"),
            ("lineEdit_12", "Actual end day"),
            ("lineEdit_13", "Supplier"),
            ("lineEdit_14", "Notes"),
            ("lineEdit_15", "Reference"),
        ]
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Metal FIN")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", f"Không thể đọc file Excel: {e}")
            return

        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Metal FIN", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from lineEdit_2 ("Name") as the filename
                image_name = self.ui.lineEdit_2.text().strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, "Thành công", "Đã thêm kim loại mới vào danh sách.")
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", f"Không thể ghi file Excel: {e}")

class WoodFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_WoodDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        self.uploaded_image_path = None

        self.ui.UpAnhGoButton.clicked.connect(self.upload_image)
        self.ui.ThemGoButton.clicked.connect(self.add_new_wood)

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Chọn ảnh JPG",
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path

    def add_new_wood(self):
        mapping = [
            ("lineEdit_1", "Code"),
            ("lineEdit_2", "Name"),
            ("lineEdit_3", "Status"),
            ("lineEdit_4", "PO"),
            ("lineEdit_5", "Generation"),
            ("lineEdit_6", "Process"),
            ("lineEdit_7", "Request day"),
            ("lineEdit_8", "Qty"),
            ("lineEdit_9", "Approved by"),
            ("lineEdit_10", "Approved day"),
            ("lineEdit_11", "Reject"),
            ("lineEdit_12", "Actual end day"),
            ("lineEdit_13", "Supplier"),
            ("lineEdit_14", "Notes"),
            ("lineEdit_15", "Master"),
            ("lineEdit_16", "Reference"),
        ]
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Wood FIN")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", f"Không thể đọc file Excel: {e}")
            return

        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Wood FIN", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from lineEdit_2 ("Name") as the filename
                image_name = self.ui.lineEdit_2.text().strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, "Thành công", "Đã thêm gỗ mới vào danh sách.")
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", f"Không thể ghi file Excel: {e}")

class BVSTDWindow(QDialog):
    def __init__(self, company_folder, server_path):
        super().__init__()
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

        self.company_folder = company_folder
        self.server_path = server_path  # <-- Add this line

        self.target_folder = os.path.join(company_folder, "TLC DRAWINGS", "STANDARD DRAWNGS  SEND TO SUPPLIER")

        # Sự kiện
        self.ui.searchLineEdit.textChanged.connect(self.perform_search)
        self.ui.resultList.itemClicked.connect(self.show_preview)
        self.ui.downloadButton.clicked.connect(self.download_file)
        self.ui.PrintButton.clicked.connect(self.print_image)
        
        self.found_files = []

    def perform_search(self):
        keyword = self.ui.searchLineEdit.text().strip().lower()
        self.ui.resultList.clear()
        self.found_files = []

        if not os.path.exists(self.target_folder):
            return

        for root, dirs, files in os.walk(self.target_folder):
            for file in files:
                if keyword in file.lower():
                    full_path = os.path.join(root, file)
                    self.found_files.append(full_path)
                    item = QListWidgetItem(file)
                    self.ui.resultList.addItem(item)

    def show_preview(self, item):
        import glob
        import difflib
        from PyQt5.QtGui import QPixmap

        # Get the base name (without extension) of the selected item
        selected_name = os.path.splitext(item.text())[0].lower()

        # Search only in Products Image folder
        products_image_dir = os.path.join(self.server_path, "Images", "Products Image")
        image_files = []
        if os.path.exists(products_image_dir):
            image_files.extend(glob.glob(os.path.join(products_image_dir, "*.jpg")))
            image_files.extend(glob.glob(os.path.join(products_image_dir, "*.jpeg")))
            image_files.extend(glob.glob(os.path.join(products_image_dir, "*.png")))

        # Find the closest match by filename (without extension)
        image_basenames = [os.path.splitext(os.path.basename(f))[0].lower() for f in image_files]
        matches = difflib.get_close_matches(selected_name, image_basenames, n=1, cutoff=0.6)

        if matches:
            # Get the full path of the best match
            best_match_index = image_basenames.index(matches[0])
            best_image_path = image_files[best_match_index]
            pixmap = QPixmap(best_image_path)
            if not pixmap.isNull():
                scaled_pixmap = pixmap.scaled(
                    self.ui.previewLabel.width(),
                    self.ui.previewLabel.height(),
                    QtCore.Qt.KeepAspectRatio,
                    QtCore.Qt.SmoothTransformation
                )
                self.ui.previewLabel.setPixmap(scaled_pixmap)
                return

        # If no match or failed to load, clear the preview
        self.ui.previewLabel.clear()

    def download_file(self):
        selected_items = self.ui.resultList.selectedItems()
        if not selected_items:
            return
        file_name = selected_items[0].text()
        source_file = next((f for f in self.found_files if os.path.basename(f) == file_name), None)

        if source_file:
            target_folder = QFileDialog.getExistingDirectory(self, "Chọn thư mục để lưu")
            if target_folder:
                shutil.copy(source_file, target_folder)
                
    def print_image(self):
        selected_items = self.ui.resultList.selectedItems()
        if not selected_items:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Vui lòng chọn một file để in.")
            return
        file_name = selected_items[0].text()
        source_file = next((f for f in self.found_files if os.path.basename(f) == file_name), None)
        if not source_file:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Không tìm thấy file.")
            return

        # Handle images
        if source_file.lower().endswith(('.png', '.jpg', '.jpeg')):
            from PyQt5.QtGui import QPixmap
            pixmap = QPixmap(source_file)
            if pixmap.isNull():
                QtWidgets.QMessageBox.warning(self, "Lỗi", "Không thể mở hình ảnh.")
                return
        # Handle PDFs
        elif source_file.lower().endswith('.pdf'):
            try:
                doc = fitz.open(source_file)
                page = doc.load_page(0)  # First page
                pix = page.get_pixmap(dpi=200)
                image_bytes = pix.tobytes("ppm")
                from PyQt5.QtGui import QImage, QPixmap
                image = QImage.fromData(image_bytes)
                pixmap = QPixmap.fromImage(image)
                if pixmap.isNull():
                    QtWidgets.QMessageBox.warning(self, "Lỗi", "Không thể mở file PDF.")
                    return
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, "Lỗi", f"Không thể mở file PDF: {e}")
                return
        else:
            QtWidgets.QMessageBox.information(self, "Thông báo", "Chỉ hỗ trợ in file hình ảnh (.jpg, .jpeg, .png) hoặc PDF.")
            return

        # Print the pixmap (image or PDF as image)
        printer = QPrinter()
        dialog = QPrintDialog(printer, self)
        if dialog.exec_() == QPrintDialog.Accepted:
            painter = QtGui.QPainter()
            if painter.begin(printer):
                page_rect = printer.pageRect()
                img_rect = pixmap.rect()
                scale = min(page_rect.width() / img_rect.width(), page_rect.height() / img_rect.height())
                x = (page_rect.width() - img_rect.width() * scale) / 2
                y = (page_rect.height() - img_rect.height() * scale) / 2
                painter.translate(x, y)
                painter.scale(scale, scale)
                painter.drawPixmap(0, 0, pixmap)
                painter.end()
        else:
            # Do nothing if print dialog is cancelled
            pass


class timsp(QtWidgets.QMainWindow):
    def __init__(self, server_path):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.server_path = server_path

        # Connect the buttons
        self.ui.TimSPButton.clicked.connect(self.search_product)
        self.ui.TCKiemtraButton.clicked.connect(self.show_hangmuc_results) 
        self.ui.lineEdit.returnPressed.connect(self.search_product)

        # QCompleter for product names
        import openpyxl
        from PyQt5.QtWidgets import QCompleter

        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        product_names = set()
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            if "Matrix" in wb.sheetnames:
                ws = wb["Matrix"]
                for row in ws.iter_rows(min_row=5, min_col=2, max_col=2):
                    cell = row[0]
                    if cell.value:
                        product_names.add(str(cell.value).strip())

        completer = QCompleter(sorted(product_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit.setCompleter(completer)

    def search_product(self):
        import openpyxl
        import os
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        from PyQt5.QtGui import QFont


        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        if not os.path.exists(excel_path):
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Không tìm thấy file List Product QC.xlsx")
            return

        keyword = self.ui.lineEdit.text().strip().lower()
        if not keyword:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Vui lòng nhập từ khóa tìm kiếm.")
            return

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "Matrix" not in wb.sheetnames:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Không tìm thấy sheet 'Matrix' trong file.")
            return
        ws = wb["Matrix"]

        # Get header from row 4, columns B to M (2 to 13)
        header = [ws.cell(row=4, column=col).value for col in range(2, 14)]

        # Search for results in column B (2), from row 5
        results = []
        for row in ws.iter_rows(min_row=5, min_col=2, max_col=2):
            cell = row[0]
            value = str(cell.value).strip().lower() if cell.value else ""
            if keyword in value:
                # Get the whole row from B to M
                row_data = [ws.cell(row=cell.row, column=col).value for col in range(2, 14)]
                results.append(row_data)

        # Set up the model for QTableView
        model = QStandardItemModel()
        model.setColumnCount(len(header))
        model.setHorizontalHeaderLabels([str(h) if h is not None else "" for h in header])

        font = QFont()
        font.setBold(True)
        font.setPointSize(12)
        for col in range(model.columnCount()):
            model.setHeaderData(col, QtCore.Qt.Horizontal, font, QtCore.Qt.FontRole)

        for row_data in results:
            items = []
            for value in row_data:
                item = QStandardItem(str(value) if value is not None else "")
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                items.append(item)
            model.appendRow(items)

        self.ui.tableView.setModel(model)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

        if not results:
            QtWidgets.QMessageBox.information(self, "Kết quả", "Không tìm thấy sản phẩm phù hợp.")
        
        self.last_matrix_results = results
        self.last_matrix_header = header

    def show_hangmuc_results(self):
        import openpyxl
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        from PyQt5.QtGui import QFont

        # Get last search results and header
        matrix_header = getattr(self, "last_matrix_header", None)
        if not matrix_header:
            QtWidgets.QMessageBox.information(self, "Thông báo", "Chưa có dữ liệu kiểm tra. Vui lòng tìm kiếm trước.")
            return

        # Get selected row in tableView
        selection_model = self.ui.tableView.selectionModel()
        if not selection_model.hasSelection():
            QtWidgets.QMessageBox.information(self, "Thông báo", "Vui lòng chọn một dòng kết quả để kiểm tra.")
            return
        selected_indexes = selection_model.selectedRows()
        if not selected_indexes:
            QtWidgets.QMessageBox.information(self, "Thông báo", "Vui lòng chọn một dòng kết quả để kiểm tra.")
            return
        selected_row = selected_indexes[0].row()

        # Get the data of the selected row
        model = self.ui.tableView.model()
        row_data = []
        for col in range(model.columnCount()):
            item = model.item(selected_row, col)
            row_data.append(item.text() if item else "")

        # Open Excel and get Hangmuc sheet
        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "Hangmuc" not in wb.sheetnames:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Không tìm thấy sheet 'Hangmuc' trong file.")
            return
        ws_hangmuc = wb["Hangmuc"]

        # Prepare results: [(header, [B, C, D, E])]
        hangmuc_results = []
        for col_idx, cell_value in enumerate(row_data):
            if str(cell_value).strip().lower() == "o":
                header_value = matrix_header[col_idx]
                # Collect ALL rows in Hangmuc where column A matches header_value
                for hangmuc_row in ws_hangmuc.iter_rows(min_row=1, min_col=1, max_col=5):
                    if str(hangmuc_row[0].value).strip() == str(header_value).strip():
                        hangmuc_row_data = [cell.value for cell in hangmuc_row[1:5]]  # B, C, D, E
                        hangmuc_results.append([header_value] + hangmuc_row_data)

        # Show in dialog
        dialog = TCKiemtraDialog(self)
        model = QStandardItemModel()
        model.setColumnCount(5)
        # Get header from row 2, columns A to E (1 to 5)
        hangmuc_header = [ws_hangmuc.cell(row=2, column=col).value for col in range(1, 6)]
        model.setHorizontalHeaderLabels([str(h) if h is not None else "" for h in hangmuc_header])

        # Set header labels and make them bold and bigger
        font = QFont()
        font.setBold(True)
        font.setPointSize(12)  # Adjust as needed

        for col in range(model.columnCount()):
            model.setHeaderData(col, QtCore.Qt.Horizontal, font, QtCore.Qt.FontRole)

        for row in hangmuc_results:
            items = []
            for value in row:
                item = QStandardItem(str(value) if value is not None else "")
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                items.append(item)
            model.appendRow(items)
        dialog.tableView.setModel(model)
        dialog.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

        dialog.exec_()

class TCKiemtraDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Hạng mục kiểm tra")
        self.resize(800, 600)
        layout = QVBoxLayout(self)
        self.tableView = QtWidgets.QTableView(self)
        layout.addWidget(self.tableView)

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = ColorSearchApp()
    window.setWindowTitle("The Lacquer Company App")
    window.show()
    sys.exit(app.exec_())
