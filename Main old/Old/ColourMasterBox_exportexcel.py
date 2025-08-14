import sys
import os
import urllib.parse
import urllib.request
import pandas as pd
from PyQt5 import QtWidgets
from PyQt5 import QtCore
from form import Ui_Form
from color_form import Ui_Form as Ui_ColorDialog
from effect_form import Ui_Form as Ui_EffectDialog
from metal_form import Ui_Form as Ui_MetalDialog
from wood_form import Ui_Form as Ui_WoodDialog
import shutil
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QDialogButtonBox, QLabel, QRadioButton, QButtonGroup, QListWidgetItem, QFileDialog
from bvstd_window import Ui_Dialog 
from timsp import Ui_MainWindow
from PyQt5.QtPrintSupport import QPrintDialog, QPrinter
from PyQt5 import QtGui
import fitz  # PyMuPDF
from PyQt5.QtCore import QTranslator, QLocale, QLibraryInfo
from TCKiemtraDialog import Ui_TCDialog as Ui_TCKiemtraDialog
from PyQt5.QtGui import QIcon
from ColorwayDialog import Ui_ColorwayDialog 
from collection import Ui_Form as Ui_CollectionDialog 

import resources_rc 

def find_company_folder():
    # List all possible company folder paths
    possible_paths = [
        r"C:\Users\Admins\The Lacquer Company\Company Files - Tài liệu",
        r"C:\Users\Admins\The Lacquer Company\Company Files - Documents"
        r"C:\Users\Admin\The Lacquer Company\Company Files - Tài liệu",
        r"C:\Users\Admin\The Lacquer Company\Company Files - Documents",
        r"C:\Users\ADMIN\The Lacquer Company\Company Files - Documents",
        r"C:\Users\ADMIN\The Lacquer Company\Company Files - Tài liệu",
        r"C:\Users\ADMINS\The Lacquer Company\Company Files - Documents",
        r"C:\Users\ADMINS\The Lacquer Company\Company Files - Tài liệu",
    ]
    for path in possible_paths:
        if os.path.exists(path):
            return path
    raise RuntimeError("Không tìm thấy thư mục công ty trên máy tính này.")

def all_words_in_text(words, text):
    return all(word in text for word in words)

class ColorSearchApp(QtWidgets.QWidget):    
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.company_folder = find_company_folder()
        self.server_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "QC FOLDER",
            "MASTER COLOR LIST QC"
        )
        serverexcel_path = os.path.join(self.server_path,"TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")

        local_path = 'TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx'

        try:
            self.df_main = pd.read_excel(serverexcel_path, sheet_name='Lacquer FIN')
            self.df_custom = pd.read_excel(serverexcel_path, sheet_name='Custom color')
            self.df_metal = pd.read_excel(serverexcel_path, sheet_name='Metal FIN')
            self.df_wood = pd.read_excel(serverexcel_path, sheet_name='Wood FIN')          
            self.df_effect = pd.read_excel(serverexcel_path, sheet_name='Effect Color Swatch Statistics')
        except Exception:
            self.df_main = pd.read_excel(local_path, sheet_name='Lacquer FIN')
            self.df_custom = pd.read_excel(local_path, sheet_name='Custom color')
            self.df_metal = pd.read_excel(local_path, sheet_name='Metal FIN')
            self.df_wood = pd.read_excel(local_path, sheet_name='Wood FIN')
            self.df_effect = pd.read_excel(local_path, sheet_name='Effect Color Swatch Statistics')

        # Strip spaces from column names for robustness
        self.df_main.columns = self.df_main.columns.str.strip()
        self.df_custom.columns = self.df_custom.columns.str.strip()
        self.df_metal.columns = self.df_metal.columns.str.strip()
        self.df_wood.columns = self.df_wood.columns.str.strip()
        self.df_effect.columns = self.df_effect.columns.str.strip()

        self.ui.pushButton.clicked.connect(self.search_color)
        self.ui.HieuUngButton.clicked.connect(self.search_effect_color)  # Connect the new button
        self.ui.ThemButton.clicked.connect(self.show_option_dialog)
        self.ui.BVSTDButton.clicked.connect(self.open_bvstd_window)
        self.ui.TCQCButton.clicked.connect(self.open_timsp_window)
        self.ui.collectionButton.clicked.connect(self.open_collection_dialog)

        self.ui.lineEdit.returnPressed.connect(self.search_color)  # Allow pressing Enter to search
    
        # After loading all DataFrames in __init__:
        self.all_product_names = set()
        for df in [self.df_main, self.df_custom, self.df_metal, self.df_wood]:
            if "Name" in df.columns:
                self.all_product_names.update(df["Name"].dropna().astype(str).tolist())

        from PyQt5.QtWidgets import QCompleter

        completer = QCompleter(sorted(self.all_product_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit.setCompleter(completer)

        # Language setup
        self.translator = QtCore.QTranslator()
        self.ui.LanguageBox.addItem(QIcon(":/images/vi.png"), "Tiếng Việt", "vi")
        self.ui.LanguageBox.addItem(QIcon(":/images/en.png"), "English", "en")
        self.ui.LanguageBox.currentIndexChanged.connect(self.change_language)
        self.ui.LanguageBox.setMinimumWidth(140)

    def change_language(self):
        lang_code = self.ui.LanguageBox.currentData()
        if lang_code:
            QtWidgets.QApplication.instance().removeTranslator(self.translator)
            if self.translator.load(f"app_{lang_code}.qm"):
                QtWidgets.QApplication.instance().installTranslator(self.translator)
            # Retranslate UI
            self.ui.retranslateUi(self)

    def open_bvstd_window(self):
        self.bvstd_window = BVSTDWindow(self.company_folder, self.server_path)
        self.bvstd_window.exec_()

    def open_timsp_window(self):
        self.timsp_window = timsp(self.server_path)
        self.timsp_window.show()

    def search_color(self):
        keyword = self.ui.lineEdit.text().strip().lower()

        # Search in 'Lacquer FIN'
        words = keyword.split()
        matched = self.df_main[self.df_main['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
        
        # If not found, search in 'Custom color'
        if matched.empty:
            #words = keyword.split()
            matched = self.df_custom[self.df_custom['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
            
        if matched.empty:
            #words = keyword.split()
            matched = self.df_metal[self.df_metal['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
        
        if matched.empty:
            #words = keyword.split()
            matched = self.df_wood[self.df_wood['Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
            if matched.empty:
                self.ui.textEdit.setText(self.tr("🔍 Không tìm thấy kết quả."))
                return

        # Add the first sentence with bold and underline
        
        html=""
        for _, row in matched.iterrows():
            name = str(row.get("Name", "")).strip()
            for col in matched.columns:
                cell = row.get(col, "")
                # If cell is a Series (shouldn't be, but just in case), get the first value
                if isinstance(cell, pd.Series):
                    cell = cell.iat[0] if not cell.empty else ""
                try:
                    is_na = pd.isna(cell)
                except Exception:
                    is_na = False
                value = "" if is_na else str(cell)
                if col.strip().lower() in ["ref image"]:
                    image_path = os.path.join(self.server_path, "Images", f"{name}.jpg")
                    if os.path.exists(image_path):
                        file_uri = urllib.parse.urljoin('file:', urllib.request.pathname2url(image_path))
                        html += f'<p><b>{col}:</b><br><img src="{file_uri}" width="200"></p>'
                    else:
                        html += f"<p><b>{col}:</b> {self.tr('Không tìm thấy ảnh')}</p>"
                else:
                    html += f"<p><b>{col}:</b> {value}</p>"
            html += "<hr>"

        self.ui.textEdit.setHtml(html)

    def search_effect_color(self):
        keyword = self.ui.lineEdit.text().strip().lower()
        # Adjust the column name below if needed
        if 'Color Name' in self.df_effect.columns:
            words = keyword.split()
            matched = self.df_effect[self.df_effect['Color Name'].str.lower().apply(lambda x: all_words_in_text(words, x) if pd.notna(x) else False)]
        else:
            # Print columns for debugging if needed
            print(self.df_effect.columns.tolist())
            self.ui.textEdit.setText(self.tr("Không tìm thấy cột 'Color Name' trong sheet hiệu ứng."))
            return

        if matched.empty:
            self.ui.textEdit.setText(self.tr("🔍 Không tìm thấy kết quả trong hiệu ứng."))
            return

        html = '<p><b><u>EFFECT COLOR</u></b></p>'
        for _, row in matched.iterrows():
            name=str(row.get("Color Name","")).strip()
            for col in matched.columns:
                cell = row.get(col, "")
                value = str(cell) if not pd.isna(cell) else ""
                if col.strip().lower() in ["ref image"]:
                    image_path=os.path.join(self.server_path,"Images",f"{name}.jpg")
                    if os.path.exists(image_path):
                        file_uri=urllib.parse.urljoin('file:',urllib.request.pathname2url(image_path))
                        html += f'<p><b>{col}:</b><br><img src="{file_uri}" width="200"></p>'
                    else:
                        html += f"<p><b>{col}:</b> ({self.tr('Không tìm thấy ảnh')})</p>"
                else:
                    html += f"<p><b>{col}:</b> {value}</p>"
            html += "<hr>"

        self.ui.textEdit.setHtml(html)

    def show_option_dialog(self):
        dialog = OptionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            selected = dialog.selected_option()
            if selected == "Color":
                color_dialog = ColorFormDialog(self)
                color_dialog.exec_()
            elif selected == "Effect":
                effect_dialog = EffectFormDialog(self)
                effect_dialog.exec_()
            elif selected == "Metal":
                metal_dialog = MetalFormDialog(self)
                metal_dialog.exec_()
            elif selected == "Wood":
                wood_dialog = WoodFormDialog(self)
                wood_dialog.exec_()
            else:
                QtWidgets.QMessageBox.information(self, self.tr("Bạn đã chọn"), f"{self.tr('Bạn đã chọn')}: {selected}")
    
    def open_collection_dialog(self):
        dialog = CollectionDialog(self)
        dialog.exec_()

class OptionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Choose an Option")
        self.resize(280, 200)  # Set a reasonable default size (width, height)
        self.setMinimumWidth(280)  # Ensure the dialog is wide enough
        layout = QVBoxLayout(self)

        label = QLabel("Choose a type:")
        layout.addWidget(label)

        self.button_group = QButtonGroup(self)
        self.radio_color = QRadioButton("Color")
        self.radio_metal = QRadioButton("Metal")
        self.radio_wood = QRadioButton("Wood")
        self.radio_effect = QRadioButton("Effect")
        self.button_group.addButton(self.radio_color)
        self.button_group.addButton(self.radio_metal)
        self.button_group.addButton(self.radio_wood)
        self.button_group.addButton(self.radio_effect)
        layout.addWidget(self.radio_color)
        layout.addWidget(self.radio_metal)
        layout.addWidget(self.radio_wood)
        layout.addWidget(self.radio_effect)

        self.radio_color.setChecked(True)  # Default selection

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_option(self):
        if self.radio_color.isChecked():
            return "Color"
        elif self.radio_metal.isChecked():
            return "Metal"
        elif self.radio_wood.isChecked():
            return "Wood"
        elif self.radio_effect.isChecked():
            return "Effect"
        return None

class ColorFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_ColorDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        
        import pandas as pd
        from PyQt5.QtWidgets import QCompleter

        # Load product names from "Lacquer FIN"
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        product_names = []
        try:
            df = pd.read_excel(excel_path, sheet_name="Lacquer FIN")
            if "Name" in df.columns:
                product_names = df["Name"].dropna().astype(str).tolist()
        except Exception as e:
            product_names = []

        completer = QCompleter(sorted(product_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit_3.setCompleter(completer)
        
        
        self.them_mau_button = self.ui.ThemMauButton
        self.them_mau_button.clicked.connect(self.add_new_color)
        self.up_anh_button = self.ui.UpAnhButton
        self.up_anh_button.clicked.connect(self.upload_image)
        self.uploaded_image_path = None  # <-- Add this line
        self.ui.CapnhatMauButton.clicked.connect(self.update_color)

    def add_new_color(self):
        # Map lineEdits to columns
        mapping = [
            ("lineEdit_1", "Collection"),
            ("lineEdit_2", "Ref-Tone Code"),
            ("lineEdit_3", "Name"),
            ("lineEdit_4", "Reference"),
            ("lineEdit_5", "Status"),
            ("lineEdit_6", "Generation"),
            ("lineEdit_7", "Process"),
            ("lineEdit_8", "Request day"),
            ("lineEdit_9", "Qty"),
            ("lineEdit_10", "Approved by "),
            ("lineEdit_11", "Approved day"),
            ("lineEdit_12", "Sup - incharge"),
            ("lineEdit_13", "Master"),
            ("lineEdit_14", "Notes"),
            ("lineEdit_15", "Applied"),
        ]

        # Read values from lineEdits
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        # Load Excel file
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Lacquer FIN")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        # Append new row
        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        # Save back to Excel
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Lacquer FIN", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from "Name" as the filename
                image_name = row_data.get("Name", "new_image").strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã thêm màu mới vào danh sách."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr("Chọn ảnh JPG"),
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path  # <-- Save the path

    def update_color(self):
        # Mapping of line edits to DataFrame columns
        mapping = {
            "lineEdit_1": "Collection",
            "lineEdit_2": "Ref-Tone Code",
            "lineEdit_3": "Name",
            "lineEdit_4": "Reference",
            "lineEdit_5": "Status",
            "lineEdit_6": "Generation",
            "lineEdit_7": "Process",
            "lineEdit_8": "Request day",
            "lineEdit_9": "Qty",
            "lineEdit_10": "Approved by ",
            "lineEdit_11": "Approved day",
            "lineEdit_12": "Sup - incharge",
            "lineEdit_13": "Master",
            "lineEdit_14": "Notes",
            "lineEdit_15": "Applied",
        }

        # Read the current values from the line edits
        row_data = {}
        for edit_name, col_name in mapping.items():
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        # Validate the "Name" field
        if not row_data["Name"]:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Tên màu không được để trống."))
            return

        # Load the existing data
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Lacquer FIN")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        # Find the row to update
        try:
            row_index = df[df["Name"] == row_data["Name"]].index[0]
        except IndexError:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Không tìm thấy màu để cập nhật."))
            return

        # Update the row
        # Only update non-empty fields
        for col, value in row_data.items():
            if value:  # Only update if not empty
                df.at[row_index, col] = value

        # Save the changes
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Lacquer FIN", index=False)

            # Copy the new image if provided
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                image_name = row_data["Name"].strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã cập nhật màu."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

class EffectFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_EffectDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        self.uploaded_image_path = None
        import pandas as pd
        from PyQt5.QtWidgets import QCompleter

        # Load product names from "Effect Color Swatch Statistics"
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        effect_names = []
        try:
            df = pd.read_excel(excel_path, sheet_name="Effect Color Swatch Statistics")
            if "Color Name" in df.columns:
                effect_names = df["Color Name"].dropna().astype(str).tolist()
        except Exception:
            effect_names = []

        completer = QCompleter(sorted(effect_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit_1.setCompleter(completer)

        self.ui.UpAnhHUButton.clicked.connect(self.upload_image)
        self.ui.ThemHUButton.clicked.connect(self.add_new_effect)
        self.ui.CapnhatHUButton.clicked.connect(self.update_effect)

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr("Chọn ảnh JPG"),
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path

    def add_new_effect(self):
        mapping = [
            ("lineEdit_1", "Color Name"),
            ("lineEdit_2", "Qty"),
            ("lineEdit_3", "Approval date"),
            ("lineEdit_4", "Note"),
        ]
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Effect Color Swatch Statistics")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Effect Color Swatch Statistics", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from lineEdit_1 as the filename
                image_name = self.ui.lineEdit_1.text().strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã thêm hiệu ứng mới vào danh sách."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

    def update_effect(self):
        mapping = {
            "lineEdit_1": "Color Name",
            "lineEdit_2": "Qty",
            "lineEdit_3": "Approval date",
            "lineEdit_4": "Note",
        }

        # Read the current values from the line edits
        row_data = {}
        for edit_name, col_name in mapping.items():
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        # Validate the "Color Name" field
        if not row_data["Color Name"]:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Tên màu không được để trống."))
            return

        # Load the existing data
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Effect Color Swatch Statistics")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        # Find the row to update
        try:
            row_index = df[df["Color Name"] == row_data["Color Name"]].index[0]
        except IndexError:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Không tìm thấy hiệu ứng để cập nhật."))
            return

        # Update the row
        # Only update non-empty fields
        for col, value in row_data.items():
            if value:  # Only update if not empty
                df.at[row_index, col] = value

        # Save the changes
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Effect Color Swatch Statistics", index=False)

            # Copy the new image if provided
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                image_name = row_data["Color Name"].strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã cập nhật hiệu ứng."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

class MetalFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_MetalDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        self.uploaded_image_path = None

        import pandas as pd
        from PyQt5.QtWidgets import QCompleter

        # Load metal names from "Metal FIN"
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        metal_names = []
        try:
            df = pd.read_excel(excel_path, sheet_name="Metal FIN")
            if "Name" in df.columns:
                metal_names = df["Name"].dropna().astype(str).tolist()
        except Exception:
            metal_names = []

        completer = QCompleter(sorted(metal_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit_2.setCompleter(completer)
        self.ui.UpAnhMeButton.clicked.connect(self.upload_image)
        self.ui.ThemMetalButton.clicked.connect(self.add_new_metal)
        self.ui.CapnhatMetalButton.clicked.connect(self.update_metal)

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr("Chọn ảnh JPG"),
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path

    def add_new_metal(self):
        mapping = [
            ("lineEdit_1", "Code"),
            ("lineEdit_2", "Name"),
            ("lineEdit_3", "Status"),
            ("lineEdit_4", "Description"),
            ("lineEdit_5", "Generation"),
            ("lineEdit_6", "Process"),
            ("lineEdit_7", "Request day"),
            ("lineEdit_8", "Qty"),
            ("lineEdit_9", "Approved by"),
            ("lineEdit_10", "Approved day"),
            ("lineEdit_11", "Reject"),
            ("lineEdit_12", "Actual end day"),
            ("lineEdit_13", "Supplier"),
            ("lineEdit_14", "Notes"),
            ("lineEdit_15", "Reference"),
        ]
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Metal FIN")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Metal FIN", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from lineEdit_2 ("Name") as the filename
                image_name = self.ui.lineEdit_2.text().strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã thêm kim loại mới vào danh sách."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

    def update_metal(self):
        # Mapping of line edits to DataFrame columns
        mapping = {
            "lineEdit_1": "Code",
            "lineEdit_2": "Name",
            "lineEdit_3": "Status",
            "lineEdit_4": "Description",
            "lineEdit_5": "Generation",
            "lineEdit_6": "Process",
            "lineEdit_7": "Request day",
            "lineEdit_8": "Qty",
            "lineEdit_9": "Approved by",
            "lineEdit_10": "Approved day",
            "lineEdit_11": "Reject",
            "lineEdit_12": "Actual end day",
            "lineEdit_13": "Supplier",
            "lineEdit_14": "Notes",
            "lineEdit_15": "Reference",
        }

        # Read the current values from the line edits
        row_data = {}
        for edit_name, col_name in mapping.items():
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text().strip()

        # Validate the "Name" field (from lineEdit_2)
        if not row_data["Name"]:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Tên kim loại không được để trống."))
            return

        # Load the existing data
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Metal FIN")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        # Find the row to update by "Name"
        try:
            row_index = df[df["Name"] == row_data["Name"]].index[0]
        except IndexError:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Không tìm thấy kim loại để cập nhật."))
            return

        # Only update non-empty fields (except "Name")
        for edit_name, col_name in mapping.items():
            if col_name == "Name":
                continue
            value = row_data[col_name]
            if value:
                df.at[row_index, col_name] = value

        # Save the changes
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Metal FIN", index=False)

            # Copy the new image if provided
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                image_name = row_data["Name"].strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã cập nhật kim loại."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

class WoodFormDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_WoodDialog()
        self.ui.setupUi(self)
        self.server_path = parent.server_path
        self.uploaded_image_path = None

        import pandas as pd
        from PyQt5.QtWidgets import QCompleter

        # Load wood names from "Wood FIN"
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        wood_names = []
        try:
            df = pd.read_excel(excel_path, sheet_name="Wood FIN")
            if "Name" in df.columns:
                wood_names = df["Name"].dropna().astype(str).tolist()
        except Exception:
            wood_names = []

        completer = QCompleter(sorted(wood_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit_2.setCompleter(completer)
        self.ui.UpAnhGoButton.clicked.connect(self.upload_image)
        self.ui.ThemGoButton.clicked.connect(self.add_new_wood)
        self.ui.CapnhatGoButton.clicked.connect(self.update_wood)

    def upload_image(self):
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            self.tr("Chọn ảnh JPG"),
            "",
            "JPEG Files (*.jpg *.jpeg);;All Files (*)",
            options=options
        )
        if file_path:
            from PIL import Image
            img = Image.open(file_path)
            width, height = img.size

            file_uri = QtCore.QUrl.fromLocalFile(file_path).toString()
            html = f'<img src="{file_uri}" width="{width}" height="{height}">'
            self.ui.textEdit_1.setHtml(html)
            self.uploaded_image_path = file_path

    def add_new_wood(self):
        mapping = [
            ("lineEdit_1", "Code"),
            ("lineEdit_2", "Name"),
            ("lineEdit_3", "Status"),
            ("lineEdit_4", "PO"),
            ("lineEdit_5", "Generation"),
            ("lineEdit_6", "Process"),
            ("lineEdit_7", "Request day"),
            ("lineEdit_8", "Qty"),
            ("lineEdit_9", "Approved by"),
            ("lineEdit_10", "Approved day"),
            ("lineEdit_11", "Reject"),
            ("lineEdit_12", "Actual end day"),
            ("lineEdit_13", "Supplier"),
            ("lineEdit_14", "Notes"),
            ("lineEdit_15", "Master"),
            ("lineEdit_16", "Reference"),
        ]
        row_data = {}
        for edit_name, col_name in mapping:
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text()

        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Wood FIN")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Wood FIN", index=False)

            # Copy uploaded image if available
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                # Use the value from lineEdit_2 ("Name") as the filename
                image_name = self.ui.lineEdit_2.text().strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã thêm gỗ mới vào danh sách."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

    def update_wood(self):
        # Mapping of line edits to DataFrame columns
        mapping = {
            "lineEdit_1": "Code",
            "lineEdit_2": "Name",
            "lineEdit_3": "Status",
            "lineEdit_4": "PO",
            "lineEdit_5": "Generation",
            "lineEdit_6": "Process",
            "lineEdit_7": "Request day",
            "lineEdit_8": "Qty",
            "lineEdit_9": "Approved by",
            "lineEdit_10": "Approved day",
            "lineEdit_11": "Reject",
            "lineEdit_12": "Actual end day",
            "lineEdit_13": "Supplier",
            "lineEdit_14": "Notes",
            "lineEdit_15": "Master",
            "lineEdit_16": "Reference",
        }

        # Read the current values from the line edits
        row_data = {}
        for edit_name, col_name in mapping.items():
            line_edit = getattr(self.ui, edit_name)
            row_data[col_name] = line_edit.text().strip()

        # Validate the "Name" field (from lineEdit_2)
        if not row_data["Name"]:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Tên gỗ không được để trống."))
            return

        # Load the existing data
        excel_path = os.path.join(self.server_path, "TLC - MASTER COLOR LIST FOR PRODUCTION.xlsx")
        try:
            df = pd.read_excel(excel_path, sheet_name="Wood FIN")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể đọc file Excel')}: {e}")
            return

        # Find the row to update by "Name"
        try:
            row_index = df[df["Name"] == row_data["Name"]].index[0]
        except IndexError:
            QtWidgets.QMessageBox.warning(self, self.tr("Cảnh báo"), self.tr("Không tìm thấy gỗ để cập nhật."))
            return

        # Only update non-empty fields (except "Name")
        for edit_name, col_name in mapping.items():
            if col_name == "Name":
                continue
            value = row_data[col_name]
            if value:
                df.at[row_index, col_name] = value

        # Save the changes
        try:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name="Wood FIN", index=False)

            # Copy the new image if provided
            if self.uploaded_image_path:
                images_dir = os.path.join(self.server_path, "Images")
                os.makedirs(images_dir, exist_ok=True)
                image_name = row_data["Name"].strip() + ".jpg"
                dest_path = os.path.join(images_dir, image_name)
                shutil.copy2(self.uploaded_image_path, dest_path)

            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã cập nhật gỗ."))
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể ghi file Excel')}: {e}")

class BVSTDWindow(QDialog):
    def __init__(self, company_folder, server_path):
        super().__init__()
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

        self.company_folder = company_folder
        self.server_path = server_path 

        self.target_folder = os.path.join(company_folder, "TLC DRAWINGS", "STANDARD DRAWNGS  SEND TO SUPPLIER")

        # Sự kiện
        self.ui.searchLineEdit.textChanged.connect(self.perform_search)
        self.ui.resultList.itemClicked.connect(self.show_preview)
        self.ui.downloadButton.clicked.connect(self.download_file)
        self.ui.PrintButton.clicked.connect(self.print_image)
        self.ui.ColorwayButton.clicked.connect(self.open_colorway_ppt)
        self.ui.OpenButton.clicked.connect(self.open_selected_file)

        self.found_files = []

    def perform_search(self):
        keyword = self.ui.searchLineEdit.text().strip().lower()
        if not keyword.startswith("standard-"):
            keyword = "standard-" + keyword
        self.ui.resultList.clear()
        self.found_files = []

        if not os.path.exists(self.target_folder):
            return

        for root, dirs, files in os.walk(self.target_folder):
            for file in files:
                if keyword in file.lower():
                    full_path = os.path.join(root, file)
                    self.found_files.append(full_path)
                    item = QListWidgetItem(file)
                    self.ui.resultList.addItem(item)

    def show_preview(self, item):
        import glob
        import difflib
        from PyQt5.QtGui import QPixmap

        # Get the base name (without extension) of the selected item
        selected_name = os.path.splitext(item.text())[0].lower()

        # Search only in Products Image folder
        products_image_dir = os.path.join(self.server_path, "Images", "Products Image")
        image_files = []
        if os.path.exists(products_image_dir):
            image_files.extend(glob.glob(os.path.join(products_image_dir, "*.jpg")))
            image_files.extend(glob.glob(os.path.join(products_image_dir, "*.jpeg")))
            image_files.extend(glob.glob(os.path.join(products_image_dir, "*.png")))

        # Find the closest match by filename (without extension)
        image_basenames = [os.path.splitext(os.path.basename(f))[0].lower() for f in image_files]
        matches = difflib.get_close_matches(selected_name, image_basenames, n=1, cutoff=0.6)

        if matches:
            # Get the full path of the best match
            best_match_index = image_basenames.index(matches[0])
            best_image_path = image_files[best_match_index]
            pixmap = QPixmap(best_image_path)
            if not pixmap.isNull():
                scaled_pixmap = pixmap.scaled(
                    self.ui.previewLabel.width(),
                    self.ui.previewLabel.height(),
                    QtCore.Qt.KeepAspectRatio,
                    QtCore.Qt.SmoothTransformation
                )
                self.ui.previewLabel.setPixmap(scaled_pixmap)
                return

        # If no match or failed to load, clear the preview
        self.ui.previewLabel.clear()

    def download_file(self):
        selected_items = self.ui.resultList.selectedItems()
        if not selected_items:
            return
        file_name = selected_items[0].text()
        source_file = next((f for f in self.found_files if os.path.basename(f) == file_name), None)

        if source_file:
            target_folder = QFileDialog.getExistingDirectory(self, "Chọn thư mục để lưu")
            if target_folder:
                shutil.copy(source_file, target_folder)
                
    def print_image(self):
        selected_items = self.ui.resultList.selectedItems()
        if not selected_items:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Vui lòng chọn một file để in."))
            return
        file_name = selected_items[0].text()
        source_file = next((f for f in self.found_files if os.path.basename(f) == file_name), None)
        if not source_file:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy file."))
            return

        # Handle images
        if source_file.lower().endswith(('.png', '.jpg', '.jpeg')):
            from PyQt5.QtGui import QPixmap
            pixmap = QPixmap(source_file)
            if pixmap.isNull():
                QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không thể mở hình ảnh."))
                return
        # Handle PDFs
        elif source_file.lower().endswith('.pdf'):
            try:
                doc = fitz.open(source_file)
                page = doc.load_page(0)  # First page
                pix = page.get_pixmap(dpi=200)
                image_bytes = pix.tobytes("ppm")
                from PyQt5.QtGui import QImage, QPixmap
                image = QImage.fromData(image_bytes)
                pixmap = QPixmap.fromImage(image)
                if pixmap.isNull():
                    QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không thể mở file PDF."))
                    return
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), f"{self.tr('Không thể mở file PDF')}: {e}")
                return
        else:
            QtWidgets.QMessageBox.information(self, self.tr("Thông báo"), self.tr("Chỉ hỗ trợ in file hình ảnh (.jpg, .jpeg, .png) hoặc PDF."))
            return

        # Print the pixmap (image or PDF as image)
        printer = QPrinter()
        dialog = QPrintDialog(printer, self)
        if dialog.exec_() == QPrintDialog.Accepted:
            painter = QtGui.QPainter()
            if painter.begin(printer):
                page_rect = printer.pageRect()
                img_rect = pixmap.rect()
                scale = min(page_rect.width() / img_rect.width(), page_rect.height() / img_rect.height())
                x = (page_rect.width() - img_rect.width() * scale) / 2
                y = (page_rect.height() - img_rect.height() * scale) / 2
                painter.translate(x, y)
                painter.scale(scale, scale)
                painter.drawPixmap(0, 0, pixmap)
                painter.end()
        else:
            # Do nothing if print dialog is cancelled
            pass
    def open_colorway_ppt(self):
        import openpyxl
        from PyQt5.QtWidgets import QMessageBox
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        import difflib

        # Get the search keyword from searchLineEdit
        keyword = self.ui.searchLineEdit.text().strip()
        if not keyword:
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Vui lòng nhập từ khóa tìm kiếm."))
            return

        # Path to Excel file
        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        if not os.path.exists(excel_path):
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy file List Product QC.xlsx"))
            return

        # Open the workbook and sheet
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "RenderColorway" not in wb.sheetnames:
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy sheet 'RenderColorway'"))
            return
        ws = wb["RenderColorway"]

        # Find all colorway values and product names
        header = [cell.value for cell in ws[1]]
        try:
            col_products = header.index("Products")
            col_colorway = header.index("Color way")
        except ValueError:
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy cột 'Products' hoặc 'Color way'"))
            return

        product_list = []
        colorway_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            product = str(row[col_products]).strip() if row[col_products] else ""
            colorway = str(row[col_colorway]).strip() if row[col_colorway] else ""
            if product and colorway:
                product_list.append(product)
                colorway_list.append(colorway)

        # Find nearest matches in Products column using difflib
        # Lowercase all product names for case-insensitive matching
        product_list_lower = [p.lower() for p in product_list]
        matches_lower = difflib.get_close_matches(keyword.lower(), product_list_lower, n=10, cutoff=0.3)
        # Map back to original case
        matches = [product_list[product_list_lower.index(m)] for m in matches_lower]
        
        if not matches:
            QMessageBox.information(self, self.tr("Kết quả"), self.tr("Không tìm thấy sản phẩm phù hợp."))
            return

        # Show dialog with product matches
        dialog = ColorwayDialog(matches, self)
        if dialog.exec_() == QDialog.Accepted:
            selected_product = dialog.get_selected_colorway()
            if not selected_product:
                QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Vui lòng chọn một sản phẩm."))
                return

            # Find the corresponding Color way for the selected product
            try:
                idx = product_list.index(selected_product)
                selected_colorway = colorway_list[idx]
            except ValueError:
                QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy Color way cho sản phẩm đã chọn."))
                return

            # Build the path to the PowerPoint file
            ppt_folder = os.path.join(
                self.company_folder,
                "TLC DRAWINGS",
                "STANDARD DRAWNGS  SEND TO SUPPLIER",
                "RENDERING COLOR WAY"
            )
            ppt_path = os.path.join(ppt_folder, f"{selected_colorway}.pptx")

            if not os.path.exists(ppt_path):
                QMessageBox.warning(self, self.tr("Lỗi"), self.tr(f"Không tìm thấy file PowerPoint: {ppt_path}"))
                return

            # Open the PowerPoint file
            try:
                os.startfile(ppt_path)  # Windows only
            except Exception as e:
                QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể mở file PowerPoint')}: {e}")
    
    def open_selected_file(self):
        selected_items = self.ui.resultList.selectedItems()
        if not selected_items:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Vui lòng chọn một file để mở."))
            return
        file_name = selected_items[0].text()
        source_file = next((f for f in self.found_files if os.path.basename(f) == file_name), None)

        if source_file and os.path.exists(source_file):
            try:
                os.startfile(source_file)  # Windows only
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể mở file')}: {e}")
        else:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy file."))

class timsp(QtWidgets.QMainWindow):
    def __init__(self, server_path):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.server_path = server_path

        # Connect the buttons
        self.ui.TimSPButton.clicked.connect(self.search_product)
        self.ui.TCKiemtraButton.clicked.connect(self.show_hangmuc_results) 
        self.ui.lineEdit.returnPressed.connect(self.search_product)

        # QCompleter for product names
        import openpyxl
        from PyQt5.QtWidgets import QCompleter

        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        product_names = set()
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            if "Matrix" in wb.sheetnames:
                ws = wb["Matrix"]
                for row in ws.iter_rows(min_row=5, min_col=2, max_col=2):
                    cell = row[0]
                    if cell.value:
                        product_names.add(str(cell.value).strip())

        completer = QCompleter(sorted(product_names), self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.ui.lineEdit.setCompleter(completer)

    def search_product(self):
        import openpyxl
        import os
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        from PyQt5.QtGui import QFont


        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        if not os.path.exists(excel_path):
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr("Could not find List Product QC.xlsx"))
            return

        keyword = self.ui.lineEdit.text().strip().lower()
        if not keyword:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Vui lòng nhập từ khóa tìm kiếm."))
            return

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "Matrix" not in wb.sheetnames:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy sheet 'Matrix' trong file."))
            return
        ws = wb["Matrix"]

        # Get header from row 4, columns B to M (2 to 13)
        header = [ws.cell(row=4, column=col).value for col in range(2, 14)]

        # Search for results in column B (2), from row 5
        results = []
        for row in ws.iter_rows(min_row=5, min_col=2, max_col=2):
            cell = row[0]
            value = str(cell.value).strip().lower() if cell.value else ""
            if keyword in value:
                # Get the whole row from B to M
                row_data = [ws.cell(row=cell.row, column=col).value for col in range(2, 14)]
                results.append(row_data)

        # Set up the model for QTableView
        model = QStandardItemModel()
        model.setColumnCount(len(header))
        model.setHorizontalHeaderLabels([str(h) if h is not None else "" for h in header])

        font = QFont()
        font.setBold(True)
        font.setPointSize(12)
        for col in range(model.columnCount()):
            model.setHeaderData(col, QtCore.Qt.Horizontal, font, QtCore.Qt.FontRole)

        for row_data in results:
            items = []
            for value in row_data:
                item = QStandardItem(str(value) if value is not None else "")
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                items.append(item)
            model.appendRow(items)

        self.ui.tableView.setModel(model)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

        if not results:
            QtWidgets.QMessageBox.information(self, self.tr("Kết quả"), self.tr("Không tìm thấy sản phẩm phù hợp."))

        self.last_matrix_results = results
        self.last_matrix_header = header

    def show_hangmuc_results(self):
        import openpyxl
        from PyQt5.QtGui import QStandardItemModel, QStandardItem, QFont

        # Get last search results and header
        matrix_header = getattr(self, "last_matrix_header", None)
        if not matrix_header:
            QtWidgets.QMessageBox.information(self, self.tr("Thông báo"), self.tr("Chưa có dữ liệu kiểm tra. Vui lòng tìm kiếm trước."))
            return

        # Get selected row in tableView
        selection_model = self.ui.tableView.selectionModel()
        if not selection_model.hasSelection():
            QtWidgets.QMessageBox.information(self, self.tr("Thông báo"), self.tr("Vui lòng chọn một dòng kết quả để kiểm tra."))
            return
        selected_indexes = selection_model.selectedRows()
        if not selected_indexes:
            QtWidgets.QMessageBox.information(self, self.tr("Thông báo"), self.tr("Vui lòng chọn một dòng kết quả để kiểm tra."))
            return
        selected_row = selected_indexes[0].row()

        # Get the data of the selected row
        model = self.ui.tableView.model()
        row_data = []
        for col in range(model.columnCount()):
            item = model.item(selected_row, col)
            row_data.append(item.text() if item else "")

        # Open Excel and get Hangmuc sheet
        excel_path = os.path.join(self.server_path, "List Product QC.xlsx")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "Hangmuc" not in wb.sheetnames:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy sheet 'Hangmuc' trong file."))
            return
        ws_hangmuc = wb["Hangmuc"]

        # Prepare results: [(header, [B, C, D, E])]
        hangmuc_results = []
        for col_idx, cell_value in enumerate(row_data):
            if str(cell_value).strip().lower() == "o":
                header_value = matrix_header[col_idx]
                # Collect ALL rows in Hangmuc where column A matches header_value
                for hangmuc_row in ws_hangmuc.iter_rows(min_row=1, min_col=1, max_col=5):
                    if str(hangmuc_row[0].value).strip() == str(header_value).strip():
                        hangmuc_row_data = [cell.value for cell in hangmuc_row[1:5]]  # B, C, D, E
                        hangmuc_results.append([header_value] + hangmuc_row_data)

        # Show in dialog
        dialog = TCKiemtraDialog(self, searched_value=self.ui.lineEdit.text())

        model = QStandardItemModel()
        model.setColumnCount(5)
        # Get header from row 2, columns A to E (1 to 5)
        hangmuc_header = [ws_hangmuc.cell(row=2, column=col).value for col in range(1, 6)]
        model.setHorizontalHeaderLabels([str(h) if h is not None else "" for h in hangmuc_header])

        font = QFont()
        font.setBold(True)
        font.setPointSize(12)
        for col in range(model.columnCount()):
            model.setHeaderData(col, QtCore.Qt.Horizontal, font, QtCore.Qt.FontRole)

        for row in hangmuc_results:
            items = []
            for value in row:
                item = QStandardItem(str(value) if value is not None else "")
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                items.append(item)
            model.appendRow(items)

        dialog.ui.tableView.setModel(model)
        dialog.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

        # Show image when a row is selected
        def show_image_for_row(row):
            if row < 0:
                dialog.ui.imageLabel.clear()
                return
            value_in_D = model.item(row, 3).text() if model.item(row, 3) else ""
            if value_in_D:
                image_path = os.path.join(self.server_path, "Images", "Defect Images", f"{value_in_D}.jpg")
                if os.path.exists(image_path):
                    pixmap = QtGui.QPixmap(image_path)
                    if not pixmap.isNull():
                        scaled = pixmap.scaled(400, 400, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
                        dialog.ui.imageLabel.setPixmap(scaled)
                        return
            dialog.ui.imageLabel.clear()

        dialog.ui.tableView.selectionModel().currentRowChanged.connect(lambda current, previous: show_image_for_row(current.row()))
        if model.rowCount() > 0:
            dialog.ui.tableView.selectRow(0)
        dialog.exec_()
        

class TCKiemtraDialog(QDialog):
    def __init__(self, parent=None, searched_value=""):
        super().__init__(parent)
        self.ui = Ui_TCKiemtraDialog()
        self.ui.setupUi(self)
        self.ui.exportexcelButton.clicked.connect(self.export_to_excel)

        # Set lineEdit_2 with the searched value from timsp
        self.ui.lineEdit_2.setText(searched_value)

    def export_to_excel(self):
        import openpyxl
        from PyQt5.QtWidgets import QMessageBox

        # Path to the template Excel file
        excel_path = os.path.join(self.parent().server_path, "QC Check list - Format.xlsx")
        if not os.path.exists(excel_path):
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy file QC Check list - Format.xlsx"))
            return

        # Open the workbook and select the correct sheet
        wb = openpyxl.load_workbook(excel_path)
        if "Format_fin_en" not in wb.sheetnames:
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không tìm thấy sheet 'Format_fin_en'"))
            return
        ws = wb["Format_fin_en"]

        # Get data from tableView
        model = self.ui.tableView.model()
        if model is None or model.rowCount() == 0:
            QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không có dữ liệu để xuất."))
            return

        # Write data to the Excel sheet
        for row in range(model.rowCount()):
            # index 0 -> column B, index 1 -> column A, index 2 -> column D, index 3 -> column G, index 4 -> column L
            ws.cell(row=9+row, column=2).value = model.item(row, 0).text() if model.item(row, 0) else ""
            ws.cell(row=9+row, column=1).value = model.item(row, 1).text() if model.item(row, 1) else ""
            ws.cell(row=9+row, column=4).value = model.item(row, 2).text() if model.item(row, 2) else ""
            ws.cell(row=9+row, column=7).value = model.item(row, 3).text() if model.item(row, 3) else ""
            ws.cell(row=9+row, column=12).value = model.item(row, 4).text() if model.item(row, 4) else ""
        
        # Assign lineEdit and lineEdit_2 values to specific cells
        ws["C4"].value = self.ui.lineEdit.text().upper()
        ws["L4"].value = self.ui.lineEdit_2.text().upper()

        # Calculate the last row you wrote to
        last_row = 9 + model.rowCount() - 1  # Data starts at row 9

        # Set the print area: columns A to L (1 to 12), rows 9 to last_row
        ws.print_area = f"A1:P{last_row}"
        
        # Save as a new file (ask user where to save)
        from PyQt5.QtWidgets import QFileDialog
        # Get values from lineEdit and lineEdit_2 for the filename
        name1 = self.ui.lineEdit.text().upper().strip()
        name2 = self.ui.lineEdit_2.text().upper().strip()
        # Sanitize filename (remove or replace invalid characters)
        import re
        def sanitize(s):
            return re.sub(r'[\\/*?:"<>|]', "_", s)
        filename = f"{sanitize(name1)}_{sanitize(name2)}.xlsx" if name1 or name2 else "QC Check list - export.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(self, self.tr("Lưu file Excel"), filename, "Excel Files (*.xlsx)")

        def sanitize1(s):
            # Excel sheet names can't have: : \ / ? * [ ]
            return re.sub(r'[:\\/*?\[\]]', "_", s)[:31]  # Excel sheet name max length is 31

        sheet_name = f"{sanitize1(name1)}_{sanitize1(name2)}" if name1 or name2 else "Sheet1"
        # Rename the worksheet
        ws.title = sheet_name

        if save_path:
            try:
                wb.save(save_path)
                QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã xuất dữ liệu ra file Excel thành công!"))
            except Exception as e:
                QMessageBox.critical(self, self.tr("Lỗi"), f"{self.tr('Không thể lưu file Excel')}: {e}")

class ColorwayDialog(QDialog):
    def __init__(self, results, parent=None):
        super().__init__(parent)
        self.ui = Ui_ColorwayDialog()
        self.ui.setupUi(self)

        # Fill the listView with results
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        self.model = QStandardItemModel()
        for product in results:
            item = QStandardItem(product)
            self.model.appendRow(item)
        self.ui.listView.setModel(self.model)

        # Connect OK and Cancel buttons
        self.ui.okButton.clicked.connect(self.accept)
        self.ui.cancelButton.clicked.connect(self.reject)

    def get_selected_colorway(self):
        indexes = self.ui.listView.selectedIndexes()
        if indexes:
            return self.model.item(indexes[0].row()).text()
        return None
    
class CollectionDialog(QDialog):
    COLLECTION_MAP = {
        "MILESREDDButton": "MILES REDD",
        "SUZANNESHARPButton": "SUZANNE SHARP",
        "VEEREGRENNEYButton": "VEERE GRENNEY",
        "PETERMIKICButton": "PETER MIKIC",
        "RITAKONIGButton": "RITA KONIG",
        "KRBButton": "KRB",
        "JANECHURCHILLButton": "JANE CHURCHILL",
        "PENTREATHHALLButton": "PENTREATH & HALL",
        "STEVENGAMBRELButton": "STEVEN GAMBREL",
        "LUKEEDWARDHALLButton": "LUKE EDWARD HALL",
        "JOHNDERIANButton": "JOHN DERIAN",
        "HOWELONDONButton": "HOWE LONDON",
        "THELACQUERCOMPANYButton": "THE LACQUER COMPANY",
        "SCHUMACHERButton": "SCHUMACHER",
        "CHRISTOPHERSPITZMILLERButton": "CHRISTOPHER SPITZMILLER",
        "SALVESENGRAHAMButton": "SALVESEN GRAHAM",
        "JEFFREYBILHUBERButton": "JEFFREY BILHUBER",
        "CAMPBELLREYButton": "CAMPBELL-REY",
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_CollectionDialog()
        self.ui.setupUi(self)

        # Load Excel data once
        import pandas as pd
        import os

        self.df = None
        try:
            excel_path = os.path.join(parent.server_path, "List Product QC.xlsx")
            self.df = pd.read_excel(excel_path, sheet_name="DATA")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Error"), self.tr(f"Could not read Excel file: {e}"))

        # Connect all buttons to the same handler
        for btn_name, collection_value in self.COLLECTION_MAP.items():
            btn = getattr(self.ui, btn_name, None)
            if btn:
                btn.clicked.connect(lambda checked, col=collection_value: self.show_collection_products(col))

    def show_collection_products(self, collection_value):
        if self.df is None:
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr("No data loaded."))
            return

        # Filter and get distinct product names
        filtered = self.df[self.df["COLLECTION"].astype(str).str.strip().str.upper() == collection_value.upper()]
        products = sorted(set(filtered["PRODUCT NAME"].dropna().astype(str)))

        # Show in tableView
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        model = QStandardItemModel()
        model.setHorizontalHeaderLabels(["PRODUCT NAME"])
        for prod in products:
            item = QStandardItem(prod)
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            model.appendRow([item])
        self.ui.tableView.setModel(model)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)

    translator = QTranslator()
    lang_code = "vi"  # or "en", or use a config/setting
    translator.load(f"app_{lang_code}.qm")
    app.installTranslator(translator)

    window = ColorSearchApp()
    window.setWindowTitle(window.tr("THE LACQUER COMPANY APP"))
    window.show()
    sys.exit(app.exec_())