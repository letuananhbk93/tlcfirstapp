import os
import sys
from weakref import proxy
import pandas as pd
from PyQt5 import QtWidgets, QtCore
from WH_form import Ui_WHForm
from nhap_form import Ui_Form as Ui_NhapForm
from xuat_form import Ui_Form as Ui_XuatForm
from diair_form import Ui_Form as Ui_DiairForm
from llcsample_form import Ui_Form as Ui_LLCSampleForm
import datetime
import openpyxl
from PyQt5.QtWidgets import QCompleter
from PyQt5.QtCore import Qt, QSortFilterProxyModel

def find_company_folder():
    possible_paths = [
        r"C:\Users\Admins\The Lacquer Company\Company Files - Tài liệu",
        r"C:\Users\Admins\The Lacquer Company\Company Files - Documents",
        r"C:\Users\Admin\The Lacquer Company\Company Files - Tài liệu",
        r"C:\Users\Admin\The Lacquer Company\Company Files - Documents",
        r"C:\Users\ADMIN\The Lacquer Company\Company Files - Documents",
        r"C:\Users\ADMIN\The Lacquer Company\Company Files - Tài liệu",
        r"C:\Users\ADMINS\The Lacquer Company\Company Files - Documents",
        r"C:\Users\ADMINS\The Lacquer Company\Company Files - Tài liệu",
    ]
    for path in possible_paths:
        if os.path.exists(path):
            return path
    raise RuntimeError("Không tìm thấy thư mục công ty trên máy tính này.")

class MultiFilterProxy(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.filter_column = -1
        self.checked_values = []
        self.keyword = ""

    def set_filter(self, column, checked_values, keyword):
        self.filter_column = column
        self.checked_values = checked_values
        self.keyword = keyword.lower()
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row, source_parent):
        if self.filter_column < 0:
            return True
        idx = self.sourceModel().index(source_row, self.filter_column, source_parent)
        data = str(self.sourceModel().data(idx))
        data_lower = data.lower()
        if self.keyword and self.keyword not in data_lower:
            return False
        # If checked_values is not None and is empty, hide all rows
        if self.checked_values is not None:
            if len(self.checked_values) == 0:
                return False
            if data not in self.checked_values:
                return False
        return True

class WarehouseForm(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_WHForm()
        self.ui.setupUi(self)
        self._current_proxy = None

        self.company_folder = find_company_folder()
        self.ui.NXTButton.clicked.connect(self.load_nxt_sheet)
        self.ui.nhapButton.clicked.connect(self.load_nhap_sheet)
        self.ui.xuatButton.clicked.connect(self.load_xuat_sheet)
        self.ui.theodoiairButton.clicked.connect(self.load_theodoiair_sheet)
        self.ui.LLCsampleButton.clicked.connect(self.load_llcsample_sheet)
        #self.ui.filterButton.clicked.connect(self.filter_by_date)
        self.ui.inputdataButton.clicked.connect(self.show_inputdata_dialog)
        self.ui.viewSKUButton.clicked.connect(self.viewSKU)
        self.ui.viewcrateButton.clicked.connect(self.viewcrate)
        self.ui.exportexcelButton.clicked.connect(self.exportexcel)

        self.current_df = None  # Store the current DataFrame
        self.current_sheet = None  # Store the current sheet name
        self.original_df = None  # Store the original DataFrame for viewcrate

    def load_nxt_sheet(self):
        # Build the Excel file path
        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )
        if not os.path.exists(excel_path):
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không tìm thấy file:\n{excel_path}"))
            return

        try:
            df = pd.read_excel(excel_path, sheet_name="N-X-T", header=None)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Error"), self.tr(f"Không thể đọc sheet 'N-X-T': {e}"))
            return

        # Set headers from the second row (index 1)
        df.columns = df.iloc[1]
        # Get data from the third row (index 2) onward
        df = df.iloc[2:].reset_index(drop=True)

        # Remove line breaks in "CRATE NAME" column if it exists
        if "CRATE NAME" in df.columns:
            df["CRATE NAME"] = df["CRATE NAME"].astype(str).str.replace('\n', '', regex=False).str.replace('\r', '', regex=False)

        # Select columns A to L (0 to 11), except column E (4)
        cols_to_show = [i for i in range(12) if i != 4]
        df = df.iloc[:, cols_to_show]

        # Show DataFrame in tableView
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        model = QStandardItemModel()
        model.setColumnCount(len(df.columns))
        model.setHorizontalHeaderLabels([str(col) for col in df.columns])

        for row in df.itertuples(index=False):
            items = [QStandardItem(str(cell) if cell is not None else "") for cell in row]
            for item in items:
                item.setTextAlignment(QtCore.Qt.AlignCenter)
            model.appendRow(items)

        # Set up proxy model for filtering
        self.original_model = model  # Store the original model
        proxy = MultiFilterProxy(self)
        proxy.setSourceModel(model)
        self._current_proxy = proxy  # <-- Set this BEFORE connecting the context menu!
        self.ui.tableView.setModel(self._current_proxy)

        header = self.ui.tableView.horizontalHeader()
        header.setSectionsClickable(True)
        header.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        try:
            header.customContextMenuRequested.disconnect()
        except Exception:
            pass
        header.customContextMenuRequested.connect(
            lambda pos: self.mo_list_Filter(header.logicalIndexAt(pos), self._current_proxy, header.mapToGlobal(pos))
        )
        self.ui.tableView.resizeColumnsToContents()
        self.ui.tableView.horizontalHeader().setStretchLastSection(True)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

        self.current_df = df.copy()
        self.current_sheet = "N-X-T"
        self.original_df = df.copy()
        self.current_df = df.copy()
        #self.show_table(self.current_df)
        self.is_sku_pivot = False
        self.update_summary_lineedits()

    def mo_list_Filter(self, col, proxy, pos):
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLineEdit, QListWidget, QListWidgetItem, QPushButton, QLabel

        if col < 0 or proxy is None:
            return

        # Only show values currently visible in the table (after filtering)
        model = self.ui.tableView.model()
        values = set()
        for row in range(model.rowCount()):
            index = model.index(row, col)
            values.add(str(model.data(index)))
        values = sorted(values)

        dlg = QDialog(self)
        dlg.setWindowTitle("Lọc cột")
        layout = QVBoxLayout(dlg)

        layout.addWidget(QLabel("Nhập từ khóa lọc:"))
        line_edit = QLineEdit()
        layout.addWidget(line_edit)

        layout.addWidget(QLabel("Chọn giá trị:"))
        list_widget = QListWidget()
        list_widget.setSelectionMode(QListWidget.MultiSelection)
        for v in values:
            item = QListWidgetItem(v)
            item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
            # Only use checked_values if this column is the filtered column
            if proxy.filter_column == col and proxy.checked_values:
                if v in proxy.checked_values:
                    item.setCheckState(QtCore.Qt.Checked)
                else:
                    item.setCheckState(QtCore.Qt.Unchecked)
            else:
                item.setCheckState(QtCore.Qt.Checked)
            list_widget.addItem(item)
        layout.addWidget(list_widget)

        btn_apply = QPushButton("Áp dụng")
        btn_clear = QPushButton("Bỏ lọc")
        btn_check_all = QPushButton("Check All")
        btn_uncheck_all = QPushButton("Uncheck All")
        layout.addWidget(btn_apply)
        layout.addWidget(btn_clear)
        layout.addWidget(btn_check_all)
        layout.addWidget(btn_uncheck_all)

        def check_all():
            for i in range(list_widget.count()):
                list_widget.item(i).setCheckState(QtCore.Qt.Checked)

        def uncheck_all():
            for i in range(list_widget.count()):
                list_widget.item(i).setCheckState(QtCore.Qt.Unchecked)

        btn_check_all.clicked.connect(check_all)
        btn_uncheck_all.clicked.connect(uncheck_all)

        def use_filter():
            checked = [list_widget.item(i).text() for i in range(list_widget.count())
                    if list_widget.item(i).checkState() == QtCore.Qt.Checked]
            keyword = line_edit.text().strip()
            proxy.set_filter(col, checked, keyword)
            self.current_df = self.get_filtered_df()  # Update to filtered data
            self.update_summary_lineedits()
            dlg.accept()

        def unuse_filter():
            proxy.set_filter(-1, [], "")
            self.current_df = self.get_filtered_df()  # Update to filtered data (should be all)
            self.update_summary_lineedits()
            dlg.accept()


        btn_apply.clicked.connect(use_filter)
        btn_clear.clicked.connect(unuse_filter)

        dlg.exec_()
        
    def load_nhap_sheet(self):
        # Build the Excel file path
        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )
        if not os.path.exists(excel_path):
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không tìm thấy file:\n{excel_path}"))
            return

        try:
            df = pd.read_excel(excel_path, sheet_name="NHẬP", header=None)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Error"), self.tr(f"Không thể đọc sheet 'NHẬP': {e}"))
            return

        # Set headers from the second row (index 1)
        df.columns = df.iloc[1]
        # Get data from the third row (index 2) onward
        df = df.iloc[2:].reset_index(drop=True)

        # Remove line breaks in "CRATE NAME" column if it exists
        if "CRATE NAME" in df.columns:
            df["CRATE NAME"] = df["CRATE NAME"].astype(str).str.replace('\n', '', regex=False).str.replace('\r', '', regex=False)

        # Select columns A to T (0 to 19), except I(8),R(17)
        exclude = [8, 17]
        cols_to_show = [i for i in range(20) if i not in exclude]
        df = df.iloc[:, cols_to_show]

        # Format the first column (date column) to dd-MM-yy if it's a datetime
        date_col = df.columns[0]
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.strftime("%d-%m-%y")

        # Show DataFrame in tableView
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        model = QStandardItemModel()
        model.setColumnCount(len(df.columns))
        model.setHorizontalHeaderLabels([str(col) for col in df.columns])

        for row in df.itertuples(index=False):
            items = [QStandardItem(str(cell) if cell is not None else "") for cell in row]
            for item in items:
                item.setTextAlignment(QtCore.Qt.AlignCenter)
            model.appendRow(items)

        # Set up proxy model for filtering
        self.original_model = model  # Store the original model
        proxy = MultiFilterProxy(self)
        proxy.setSourceModel(model)
        self._current_proxy = proxy  # <-- Set this BEFORE connecting the context menu!
        self.ui.tableView.setModel(self._current_proxy)

        # Set up custom filter header
        header = self.ui.tableView.horizontalHeader()
        header.setSectionsClickable(True)
        header.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        try:
            header.customContextMenuRequested.disconnect()
        except Exception:
            pass
        header.customContextMenuRequested.connect(
            lambda pos: self.mo_list_Filter(header.logicalIndexAt(pos), self._current_proxy, header.mapToGlobal(pos))
        )

        self.ui.tableView.resizeColumnsToContents()
        self.ui.tableView.horizontalHeader().setStretchLastSection(True)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

        self.current_df = df.copy()
        self.current_sheet = "NHẬP"
        self.original_df = df.copy()
        self.current_df = df.copy()
        #self.show_table(self.current_df)
        self.is_sku_pivot = False
        self.is_sku_pivot_NXT = False
        self.update_summary_lineedits()

    def load_xuat_sheet(self):
        # Build the Excel file path
        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )
        if not os.path.exists(excel_path):
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không tìm thấy file:\n{excel_path}"))
            return

        try:
            df = pd.read_excel(excel_path, sheet_name="XUẤT", header=None)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Error"), self.tr(f"Không thể đọc sheet 'XUẤT': {e}"))
            return

        # Set headers from the second row (index 1)
        df.columns = df.iloc[1]
        # Get data from the third row (index 2) onward
        df = df.iloc[2:].reset_index(drop=True)

        # Remove line breaks in "CRATE NAME" column if it exists
        if "CRATE NAME" in df.columns:
            df["CRATE NAME"] = df["CRATE NAME"].astype(str).str.replace('\n', '', regex=False).str.replace('\r', '', regex=False)

        # Select columns A to S (0 to 18), except G(6), K(10), L(11), M(12), N(13), O(14), P(15), Q(16), R(17)
        exclude = [6, 10, 11, 12, 13, 14, 15, 16, 17]
        cols_to_show = [i for i in range(19) if i not in exclude]
        df = df.iloc[:, cols_to_show]

        # Format the first column (date column) to dd-MM-yy if it's a datetime
        date_col = df.columns[0]
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.strftime("%d-%m-%y")

        # Show DataFrame in tableView
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        model = QStandardItemModel()
        model.setColumnCount(len(df.columns))
        model.setHorizontalHeaderLabels([str(col) for col in df.columns])

        for row in df.itertuples(index=False):
            items = [QStandardItem(str(cell) if cell is not None else "") for cell in row]
            for item in items:
                item.setTextAlignment(QtCore.Qt.AlignCenter)
            model.appendRow(items)

        # Set up proxy model for filtering
        self.original_model = model  # Store the original model
        proxy = MultiFilterProxy(self)
        proxy.setSourceModel(model)
        self._current_proxy = proxy  # <-- Set this BEFORE connecting the context menu!
        self.ui.tableView.setModel(self._current_proxy)

        # Set up custom filter header
        header = self.ui.tableView.horizontalHeader()
        header.setSectionsClickable(True)
        header.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        try:
            header.customContextMenuRequested.disconnect()
        except Exception:
            pass
        header.customContextMenuRequested.connect(
            lambda pos: self.mo_list_Filter(header.logicalIndexAt(pos), self._current_proxy, header.mapToGlobal(pos))
        )
        
        self.ui.tableView.resizeColumnsToContents()
        self.ui.tableView.horizontalHeader().setStretchLastSection(True)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

        self.current_df = df.copy()
        self.current_sheet = "XUẤT"
        self.original_df = df.copy()
        self.current_df = df.copy()
        #self.show_table(self.current_df)
        self.is_sku_pivot = False
        self.is_sku_pivot_NXT = False
        # Update summary text edits
        self.update_summary_lineedits()

    def load_theodoiair_sheet(self):
        # Build the Excel file path
        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )
        if not os.path.exists(excel_path):
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không tìm thấy file:\n{excel_path}"))
            return

        try:
            df = pd.read_excel(excel_path, sheet_name="THEO DÕI HÀNG ĐI AIR", header=None)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Error"), self.tr(f"Không thể đọc sheet 'THEO DÕI HÀNG ĐI AIR': {e}"))
            return

        # Set headers from the second row (index 0)
        df.columns = df.iloc[0]
        # Get data from the third row (index 1) onward
        df = df.iloc[1:].reset_index(drop=True)

        # Remove line breaks in "CRATE NAME" column if it exists
        if "CRATE NAME" in df.columns:
            df["CRATE NAME"] = df["CRATE NAME"].astype(str).str.replace('\n', '', regex=False).str.replace('\r', '', regex=False)

        # Select columns A to S (0 to 18), except I (8)
        exclude = [8]
        cols_to_show = [i for i in range(19) if i not in exclude]
        df = df.iloc[:, cols_to_show]

        # Format the first column (date column) to dd-MM-yy if it's a datetime
        date_col = df.columns[0]
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.strftime("%d-%m-%y")

        # Show DataFrame in tableView
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        model = QStandardItemModel()
        model.setColumnCount(len(df.columns))
        model.setHorizontalHeaderLabels([str(col) for col in df.columns])

        for row in df.itertuples(index=False):
            items = [QStandardItem(str(cell) if cell is not None else "") for cell in row]
            for item in items:
                item.setTextAlignment(QtCore.Qt.AlignCenter)
            model.appendRow(items)

        # Set up proxy model for filtering
        self.original_model = model  # Store the original model
        proxy = MultiFilterProxy(self)
        proxy.setSourceModel(model)
        self._current_proxy = proxy  # <-- Set this BEFORE connecting the context menu!
        self.ui.tableView.setModel(self._current_proxy)

        # Set up custom filter header
        header = self.ui.tableView.horizontalHeader()
        header.setSectionsClickable(True)
        header.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        try:
            header.customContextMenuRequested.disconnect()
        except Exception:
            pass
        header.customContextMenuRequested.connect(
            lambda pos: self.mo_list_Filter(header.logicalIndexAt(pos), self._current_proxy, header.mapToGlobal(pos))
        )
        
        self.ui.tableView.resizeColumnsToContents()
        self.ui.tableView.horizontalHeader().setStretchLastSection(True)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
        
        self.current_df = df.copy()
        self.current_sheet = "THEO DÕI HÀNG ĐI AIR"
        self.original_df = df.copy()
        self.current_df = df.copy()
        #self.show_table(self.current_df)
        self.is_sku_pivot = False
        self.is_sku_pivot_NXT = False
        self.update_summary_lineedits()
        
    def load_llcsample_sheet(self):
        # Build the Excel file path
        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )
        if not os.path.exists(excel_path):
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không tìm thấy file:\n{excel_path}"))
            return

        try:
            df = pd.read_excel(excel_path, sheet_name="LLC SAMPLE", header=None)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Error"), self.tr(f"Không thể đọc sheet 'LLC SAMPLE': {e}"))
            return

        # Set headers from the second row (index 0)
        df.columns = df.iloc[0]
        # Get data from the third row (index 1) onward
        df = df.iloc[1:].reset_index(drop=True)

        # Remove line breaks in "CRATE NAME" column if it exists
        if "CRATE NAME" in df.columns:
            df["CRATE NAME"] = df["CRATE NAME"].astype(str).str.replace('\n', '', regex=False).str.replace('\r', '', regex=False)

        # Select columns A to N (0 to 13)
        cols_to_show = list(range(14))
        df = df.iloc[:, cols_to_show]

        # Format the first column (date column) to dd-MM-yy if it's a datetime
        date_col = df.columns[0]
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.strftime("%d-%m-%y")

        # Show DataFrame in tableView
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        model = QStandardItemModel()
        model.setColumnCount(len(df.columns))
        model.setHorizontalHeaderLabels([str(col) for col in df.columns])

        for row in df.itertuples(index=False):
            items = [QStandardItem(str(cell) if cell is not None else "") for cell in row]
            for item in items:
                item.setTextAlignment(QtCore.Qt.AlignCenter)
            model.appendRow(items)

        # Set up proxy model for filtering
        self.original_model = model  # Store the original model
        proxy = MultiFilterProxy(self)
        proxy.setSourceModel(model)
        self._current_proxy = proxy
        self.ui.tableView.setModel(proxy)

        # Set up custom filter header
        header = self.ui.tableView.horizontalHeader()
        header.setSectionsClickable(True)
        header.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        try:
            header.customContextMenuRequested.disconnect()
        except Exception:
            pass
        header.customContextMenuRequested.connect(
            lambda pos: self.mo_list_Filter(header.logicalIndexAt(pos), self._current_proxy, header.mapToGlobal(pos))
        )

        self.ui.tableView.resizeColumnsToContents()
        self.ui.tableView.horizontalHeader().setStretchLastSection(True)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

        self.current_df = df.copy()
        self.current_sheet = "LLC SAMPLE"
        self.original_df = df.copy()
        self.current_df = df.copy()
        #self.show_table(self.current_df)
        self.is_sku_pivot = False
        self.is_sku_pivot_NXT = False
        self.update_summary_lineedits()

    def show_inputdata_dialog(self):
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(self.tr("Chọn loại dữ liệu nhập"))
        dialog.resize(300, 200)
        layout = QtWidgets.QVBoxLayout(dialog)

        btn_nhap = QtWidgets.QPushButton(self.tr("NHẬP"))
        btn_xuat = QtWidgets.QPushButton(self.tr("XUẤT"))
        btn_diair = QtWidgets.QPushButton(self.tr("THEO DÕI HÀNG ĐI AIR"))
        btn_llcsample = QtWidgets.QPushButton(self.tr("LLC SAMPLE"))

        layout.addWidget(btn_nhap)
        layout.addWidget(btn_xuat)
        layout.addWidget(btn_diair)
        layout.addWidget(btn_llcsample)

        btn_nhap.clicked.connect(lambda: self.open_nhap_dialog(dialog))
        btn_xuat.clicked.connect(lambda: self.open_xuat_dialog(dialog))
        btn_diair.clicked.connect(lambda: self.open_diair_dialog(dialog))
        btn_llcsample.clicked.connect(lambda: self.open_llcsample_dialog(dialog))

        dialog.exec_()
    
    def show_table(self, df):
        from PyQt5.QtGui import QStandardItemModel, QStandardItem
        model = QStandardItemModel()
        model.setColumnCount(len(df.columns))
        model.setHorizontalHeaderLabels([str(col) for col in df.columns])
        for row in df.itertuples(index=False):
            items = [QStandardItem(str(cell) if cell is not None else "") for cell in row]
            for item in items:
                item.setTextAlignment(QtCore.Qt.AlignCenter)
            model.appendRow(items)
        self.ui.tableView.setModel(model)
        self.ui.tableView.resizeColumnsToContents()
        self.ui.tableView.horizontalHeader().setStretchLastSection(True)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)

    def open_nhap_dialog(self, parent_dialog):
        parent_dialog.accept()
        dlg = NhapFormDialog(self)
        dlg.exec_()

    def open_xuat_dialog(self, parent_dialog):
        parent_dialog.accept()
        dlg = XuatFormDialog(self)
        dlg.exec_()

    def open_diair_dialog(self, parent_dialog):
        parent_dialog.accept()
        dlg = DiairFormDialog(self)
        dlg.exec_()

    def open_llcsample_dialog(self, parent_dialog):
        parent_dialog.accept()
        dlg = LLCSampleFormDialog(self)
        dlg.exec_()
        
    def viewSKU(self):
        import pandas as pd
        from PyQt5.QtGui import QStandardItemModel, QStandardItem

        if self.current_df is None:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Chưa có dữ liệu để xem SKU."))
            return
        if self.is_sku_pivot == True:
            return
        
        df_filtered = self.current_df.copy()

        # If current sheet is N-X-T, do a multi-column pivot
        if self.current_sheet == "N-X-T":
            sku_col = "SKU"
            so_kien_col = "CRATE NAME"
            qty_col1 = "IMPORT"  
            qty_col2 = "EXPORT"  
            qty_col3 = "BONDED WH"  
            qty_col4 = "ALLOCATED"  
            qty_col5 = "AVAILABLE"  

            # Pivot-like summary: group by SKU, sum quantity, sum packages
            summary = (
                df_filtered
                .groupby(sku_col, as_index=False)
                .agg({
                    so_kien_col: lambda x: ', '.join(
                        sorted(
                            set(
                                str(i).replace('\n', '').replace('\r', '') for i in x if pd.notna(i)
                            )
                        )
                    ),
                    qty_col1: "sum",
                    qty_col2: "sum",
                    qty_col3: "sum",
                    qty_col4: "sum",
                    qty_col5: "sum"
                })
                .rename(columns={
                    so_kien_col: "CRATE NAME",
                    qty_col1: "IMPORT",
                    qty_col2: "EXPORT",
                    qty_col3: "BONDED WH",
                    qty_col4: "ALLOCATED",
                    qty_col5: "AVAILABLE"
                })
            )

            # Prepare model for QTableView
            model = QStandardItemModel()
            model.setColumnCount(6)
            model.setHorizontalHeaderLabels(["SKU", "CRATE NAME", "IMPORT", "EXPORT", "BONDED WH", "ALLOCATED", "AVAILABLE"])

            for _, row in summary.iterrows():
                items = [
                    QStandardItem(str(row["SKU"])),
                    QStandardItem(str(row["CRATE NAME"])),
                    QStandardItem(str(row["IMPORT"])),
                    QStandardItem(str(row["EXPORT"])),
                    QStandardItem(str(row["BONDED WH"])),
                    QStandardItem(str(row["ALLOCATED"])),
                    QStandardItem(str(row["AVAILABLE"]))
                ]
                for item in items:
                    item.setTextAlignment(QtCore.Qt.AlignCenter)
                model.appendRow(items)


            self.ui.tableView.setModel(model)
            self.ui.tableView.resizeColumnsToContents()
            self.ui.tableView.horizontalHeader().setStretchLastSection(True)
            self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
            self.current_df = summary
            self.show_table(self.current_df)
            self.is_sku_pivot = True
            self.is_sku_pivot_NXT = True
            self.update_summary_lineedits()
            return

        # --- Default logic for other sheets ---
        # Assume "SKU" and "CRATE NAME" columns exist (adjust names if needed)
        sku_col = "SKU"
        so_kien_col = "CRATE NAME"
        qty_col = "QTY"  # Adjust to your actual quantity column name

        # Pivot-like summary: group by SKU, sum quantity, sum packages
        summary = (
            df_filtered
            .groupby(sku_col, as_index=False)
            .agg({
                    so_kien_col: lambda x: ', '.join(
                        sorted(
                            set(
                                str(i).replace('\n', '').replace('\r', '') for i in x if pd.notna(i)
                            )
                        )
                    ),
                    qty_col: "sum"})
            .rename(columns={so_kien_col: "Tổng Số Kiện", qty_col: "QTY"})
        )

        # Prepare model for QTableView
        model = QStandardItemModel()
        model.setColumnCount(3)
        model.setHorizontalHeaderLabels(["SKU", "Tổng Số Kiện", "QTY"])

        for _, row in summary.iterrows():
            items = [
                QStandardItem(str(row["SKU"])),
                QStandardItem(str(row["Tổng Số Kiện"])),
                QStandardItem(str(row["QTY"]))
            ]
            for item in items:
                item.setTextAlignment(QtCore.Qt.AlignCenter)
            model.appendRow(items)

        self.ui.tableView.setModel(model)
        self.ui.tableView.resizeColumnsToContents()
        self.ui.tableView.horizontalHeader().setStretchLastSection(True)
        self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
        self.current_df = summary
        self.show_table(self.current_df)
        self.is_sku_pivot = True
        self.update_summary_lineedits()

    def viewcrate(self):
        from PyQt5.QtGui import QStandardItemModel, QStandardItem

        # Use the original DataFrame
        if hasattr(self, "original_df") and self.original_df is not None:
            self.current_df = self.original_df.copy()

            # Build the model from the original DataFrame
            df = self.original_df.copy()
            model = QStandardItemModel()
            model.setColumnCount(len(df.columns))
            model.setHorizontalHeaderLabels([str(col) for col in df.columns])
            for row in df.itertuples(index=False):
                items = [QStandardItem(str(cell) if cell is not None else "") for cell in row]
                for item in items:
                    item.setTextAlignment(QtCore.Qt.AlignCenter)
                model.appendRow(items)

            # Remove proxy and show plain model (clear all filters)
            self.ui.tableView.setModel(model)
            self._current_proxy = None
            # After showing the plain model
            proxy = MultiFilterProxy(self)
            proxy.setSourceModel(model)
            self._current_proxy = proxy
            self.ui.tableView.setModel(self._current_proxy)

            header = self.ui.tableView.horizontalHeader()
            header.setSectionsClickable(True)
            header.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
            try:
                header.customContextMenuRequested.disconnect()
            except Exception:
                pass
            header.customContextMenuRequested.connect(
                lambda pos: self.mo_list_Filter(header.logicalIndexAt(pos), self._current_proxy, header.mapToGlobal(pos))
            )

            self.ui.tableView.resizeColumnsToContents()
            self.ui.tableView.horizontalHeader().setStretchLastSection(True)
            self.ui.tableView.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
            self.is_sku_pivot = False
            self.update_summary_lineedits()
        else:
            QtWidgets.QMessageBox.warning(self, self.tr("Lỗi"), self.tr("Không có dữ liệu gốc để hiển thị."))
    
    def exportexcel(self):
        import openpyxl
        import pandas as pd
        from PyQt5.QtWidgets import QFileDialog

        # Ask user where to save
        save_path, _ = QFileDialog.getSaveFileName(self, "Export Excel", "", "Excel Files (*.xlsx)")
        if not save_path:
            return

        # Check if you are in SKU pivot mode (for example, by a flag)
        if getattr(self, "is_sku_pivot", False):
            # Export only the pivot table (self.current_df should be the pivot DataFrame)
            try:
                self.current_df.to_excel(save_path, index=False)
                QtWidgets.QMessageBox.information(self, "Thành công", "Đã xuất file Excel (pivot table) thành công!")
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, "Lỗi", f"Không thể xuất file Excel:\n{e}")
        else:
            # Export the base file (copy the original Excel file)
            try:
                import shutil
                base_path = os.path.join(
                    self.company_folder,
                    "THE LACQUER COMPANY - VIETNAM OFFICE",
                    "WAREHOUSE - IN OUT",
                    "00. INVENTORY REPORT",
                    "VIETNAM STOCK LIST UPDATE-APP.xlsx"
                )
                shutil.copyfile(base_path, save_path)
                QtWidgets.QMessageBox.information(self, "Thành công", "Đã xuất file Excel gốc thành công!")
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, "Lỗi", f"Không thể xuất file Excel:\n{e}")

    def update_summary_lineedits(self):
        # Remove any existing string values in all columns that should be numeric
        for col in ["IMPORT", "EXPORT", "BONDED WH", "ALLOCATED", "AVAILABLE", "QTY"]:
                if col in self.current_df.columns:
                    self.current_df[col] = self.current_df[col].apply(
                        lambda x: x if isinstance(x, (int, float)) else None
                    )
        # N-X-T sheet in viewSKU mode
        if getattr(self, "is_sku_pivot_NXT", False) or self.current_sheet == "N-X-T":
            for colname, edit in [
                ("IMPORT", self.ui.lineEdit_1),
                ("EXPORT", self.ui.lineEdit_2),
                ("BONDED WH", self.ui.lineEdit_3),
                ("ALLOCATED", self.ui.lineEdit_4),
                ("AVAILABLE", self.ui.lineEdit_5),
            ]:
                try:
                    total = pd.to_numeric(self.current_df[colname], errors="coerce").sum()
                    edit.setText(str(int(total)))
                except Exception:
                    edit.setText("")
            self.ui.lineEdit_6.setText("")
        else:
            # Other sheets: only lineEdit_6 shows sum of QTY
            for i in range(1, 6):
                getattr(self.ui, f"lineEdit_{i}").setText("")
            try:
                total_qty = pd.to_numeric(self.current_df["QTY"], errors="coerce").sum()
                self.ui.lineEdit_6.setText(str(int(total_qty)))
            except Exception:
                self.ui.lineEdit_6.setText("")

    def get_filtered_df(self):
        """Return a DataFrame of the currently visible (filtered) rows, preserving original dtypes."""
        model = self.ui.tableView.model()
        if not isinstance(model, QSortFilterProxyModel):
            return self.current_df
        # Always use the original DataFrame for correct indexing
        if not hasattr(self, "original_df") or self.original_df is None:
            return None
        source_df = self.original_df
        # Get the original DataFrame's index for each visible row
        visible_indices = [
            model.mapToSource(model.index(row, 0)).row()
            for row in range(model.rowCount())
        ]
        # Use iloc to preserve dtypes and numeric columns
        filtered_df = source_df.iloc[visible_indices].reset_index(drop=True)
        return filtered_df

class NhapFormDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_NhapForm()
        self.ui.setupUi(self)

        self.company_folder = find_company_folder()

        import pandas as pd

        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )

        completer_list = []
        if os.path.exists(excel_path):
            try:
                df = pd.read_excel(excel_path, sheet_name="NHẬP", usecols="A", header=None)

                def format_date(val):
                    if pd.isna(val):
                        return None
                    # Check for pandas Timestamp or Python datetime
                    if isinstance(val, (pd.Timestamp, datetime.datetime)):
                        return val.strftime("%d-%m-%y")
                    # Try to parse string as date
                    try:
                        parsed = pd.to_datetime(val, errors='raise')
                        return parsed.strftime("%d-%m-%y")
                    except Exception:
                        return str(val).strip()

                completer_list = df[0].dropna().map(format_date).dropna().unique().tolist()
            except Exception as e:
                print(self.tr("Error reading Excel for completer:"), e)

        from PyQt5.QtWidgets import QCompleter
        completer = QCompleter(completer_list)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        self.ui.lineEdit_1.setCompleter(completer)
        self.ui.nhapkhoButton.clicked.connect(self.nhapkhosheet)

        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb["NHẬP"]

            def get_completer_list(col_idx):
                # col_idx: 1-based index (A=1, B=2, ...)
                return list({
                    str(row[0].value).strip()
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx)
                    if row[0].value
                })

            # Set completers for each lineEdit
            completer2 = QCompleter(get_completer_list(2))
            completer2.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer2.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_2.setCompleter(completer2)

            completer3 = QCompleter(get_completer_list(5))
            completer3.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer3.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_3.setCompleter(completer3)

            completer4 = QCompleter(get_completer_list(6))
            completer4.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer4.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_4.setCompleter(completer4)

            completer5 = QCompleter(get_completer_list(7))
            completer5.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer5.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_5.setCompleter(completer5)

            completer6 = QCompleter(get_completer_list(8))
            completer6.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer6.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_6.setCompleter(completer6)

            completer14 = QCompleter(get_completer_list(19))
            completer14.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer14.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_14.setCompleter(completer14)

            wb.close()

    def nhapkhosheet(self):
        import openpyxl

        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )

        try:
            wb = openpyxl.load_workbook(excel_path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không thể mở file Excel:\n{e}"))
            return

        # --- Append to NHẬP sheet ---
        try:
            ws_nhap = wb["NHẬP"]
            next_row_nhap = ws_nhap.max_row + 1

            # Write values to columns as specified
            ws_nhap.cell(row=next_row_nhap, column=1, value=self.ui.lineEdit_1.text())   # A
            ws_nhap.cell(row=next_row_nhap, column=2, value=self.ui.lineEdit_2.text())   # B
            ws_nhap.cell(row=next_row_nhap, column=5, value=self.ui.lineEdit_3.text())   # E
            ws_nhap.cell(row=next_row_nhap, column=6, value=self.ui.lineEdit_4.text())   # F
            ws_nhap.cell(row=next_row_nhap, column=7, value=self.ui.lineEdit_5.text())   # G
            ws_nhap.cell(row=next_row_nhap, column=8, value=self.ui.lineEdit_6.text())   # H
            ws_nhap.cell(row=next_row_nhap, column=9, value=self.ui.lineEdit_7.text())   # I
            ws_nhap.cell(row=next_row_nhap, column=11, value=self.ui.lineEdit_8.text())  # K
            ws_nhap.cell(row=next_row_nhap, column=12, value=self.ui.lineEdit_9.text())  # L
            ws_nhap.cell(row=next_row_nhap, column=13, value=self.ui.lineEdit_10.text())  # M
            ws_nhap.cell(row=next_row_nhap, column=14, value=self.ui.lineEdit_11.text()) # N
            ws_nhap.cell(row=next_row_nhap, column=15, value=self.ui.lineEdit_12.text()) # O
            ws_nhap.cell(row=next_row_nhap, column=16, value=None)                       # P(formula set below)
            ws_nhap.cell(row=next_row_nhap, column=19, value=self.ui.lineEdit_14.text()) # S
            ws_nhap.cell(row=next_row_nhap, column=17, value=self.ui.lineEdit_15.text()) # Q

            # Set formula for column P: P = L * M * N / 1000000 (columns 12, 13, 14)
            ws_nhap.cell(row=next_row_nhap, column=16, value=f"=L{next_row_nhap}*M{next_row_nhap}*N{next_row_nhap}/1000000")

        except Exception as e:
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không thể ghi sheet 'NHẬP':\n{e}"))
            wb.close()
            return

        # --- Append to N-X-T sheet ---
        try:
            ws_nxt = wb["N-X-T"]
            next_row_nxt = ws_nxt.max_row + 1

            ws_nxt.cell(row=next_row_nxt, column=1, value=self.ui.lineEdit_2.text())    # A
            ws_nxt.cell(row=next_row_nxt, column=2, value=self.ui.lineEdit_3.text())    # B
            ws_nxt.cell(row=next_row_nxt, column=3, value=self.ui.lineEdit_4.text())    # C
            ws_nxt.cell(row=next_row_nxt, column=4, value=self.ui.lineEdit_6.text())    # D
            ws_nxt.cell(row=next_row_nxt, column=6, value=self.ui.lineEdit_7.text())    # F
            ws_nxt.cell(row=next_row_nxt, column=8, value=None)                         # H (formula set below)
            ws_nxt.cell(row=next_row_nxt, column=11, value=self.ui.lineEdit_14.text())  # K
            ws_nxt.cell(row=next_row_nxt, column=10, value=None)                        # J (formula set below)
            ws_nxt.cell(row=next_row_nxt, column=9, value=None)                         # I (formula set below)

            # Set formula for column H in N-X-T if needed (example: H = F - G)
            ws_nxt.cell(row=next_row_nxt, column=8, value=f"=F{next_row_nxt}-G{next_row_nxt}")

            # Set formula for column J in N-X-T if needed (example: H = F - G)
            ws_nxt.cell(row=next_row_nxt, column=10, value=f"=F{next_row_nxt}-G{next_row_nxt}-I{next_row_nxt}")

            # Set formula for column I (9) in N-X-T
            ws_nxt.cell(
                row=next_row_nxt,
                column=9,
                value=f'=SUMIFS(XUẤT!J:J,XUẤT!F:F,\'N-X-T\'!B{next_row_nxt},XUẤT!A:A,"ALLOCATED",XUẤT!C:C,\'N-X-T\'!A{next_row_nxt})'
            )
            for i in range(1, 15):
                getattr(self.ui, f"lineEdit_{i}").setText("")

        except Exception as e:
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không thể ghi sheet 'N-X-T':\n{e}"))
            wb.close()
            return

        # --- Save workbook ---
        try:
            wb.save(excel_path)
            wb.close()
            QtWidgets.QMessageBox.information(self, "Thành công", "Đã nhập kho thành công!")
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Không thể lưu file Excel:\n{e}")

class XuatFormDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_XuatForm()
        self.ui.setupUi(self)
        self.company_folder = find_company_folder()

        # Path to your Excel file
        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )

        # --- QCompleter for lineEdit_1 from column A ---
        completer_list = []
        if os.path.exists(excel_path):
            try:
                df = pd.read_excel(excel_path, sheet_name="XUẤT", usecols="A", header=None)
                def format_date(val):
                    if pd.isna(val):
                        return None
                    if isinstance(val, (pd.Timestamp, datetime.datetime)):
                        return val.strftime("%d-%m-%y")
                    try:
                        parsed = pd.to_datetime(val, errors='raise')
                        return parsed.strftime("%d-%m-%y")
                    except Exception:
                        return str(val).strip()
                completer_list = df[0].dropna().map(format_date).dropna().unique().tolist()
            except Exception as e:
                print("Error reading Excel for completer:", e)

        from PyQt5.QtWidgets import QCompleter
        completer = QCompleter(completer_list)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        self.ui.lineEdit_1.setCompleter(completer)

        # --- Connect button ---        
        self.ui.xuatkhoButton.clicked.connect(self.xuatkhosheet)
        
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb["XUẤT"]

            def get_completer_list(col_idx):
                # col_idx: 1-based index (A=1, B=2, ...)
                return list({
                    str(row[0].value).strip()
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx)
                    if row[0].value
                })

            # Set completers for each lineEdit
            completer2 = QCompleter(get_completer_list(2))
            completer2.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer2.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_2.setCompleter(completer2)

            completer3 = QCompleter(get_completer_list(3))
            completer3.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer3.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_3.setCompleter(completer3)

            completer4 = QCompleter(get_completer_list(4))
            completer4.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer4.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_4.setCompleter(completer4)

            completer5 = QCompleter(get_completer_list(5))
            completer5.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer5.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_5.setCompleter(completer5)

            completer6 = QCompleter(get_completer_list(6))
            completer6.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer6.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_6.setCompleter(completer6)

            completer7 = QCompleter(get_completer_list(8))
            completer7.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer7.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_7.setCompleter(completer7)

            completer8 = QCompleter(get_completer_list(9))
            completer8.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer8.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_8.setCompleter(completer8)

            wb.close()

    def xuatkhosheet(self):
        import openpyxl

        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )

        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["XUẤT"]
            next_row = ws.max_row + 1

            ws.cell(row=next_row, column=1, value=self.ui.lineEdit_1.text())   # A
            ws.cell(row=next_row, column=2, value=self.ui.lineEdit_2.text())   # B
            ws.cell(row=next_row, column=3, value=self.ui.lineEdit_3.text())   # C
            ws.cell(row=next_row, column=4, value=self.ui.lineEdit_4.text())   # D
            ws.cell(row=next_row, column=5, value=self.ui.lineEdit_5.text())   # E
            ws.cell(row=next_row, column=6, value=self.ui.lineEdit_6.text())   # F
            ws.cell(row=next_row, column=8, value=self.ui.lineEdit_7.text())   # H
            ws.cell(row=next_row, column=9, value=self.ui.lineEdit_8.text())   # I
            ws.cell(row=next_row, column=10, value=self.ui.lineEdit_9.text())  # J
            ws.cell(row=next_row, column=19, value=self.ui.lineEdit_10.text()) # S

            wb.save(excel_path)
            wb.close()
            QtWidgets.QMessageBox.information(self, "Thành công", "Đã xuất kho thành công!")
            for i in range(1, 10):
                getattr(self.ui, f"lineEdit_{i}").setText("")
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Không thể ghi file Excel:\n{e}")

class DiairFormDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_DiairForm()
        self.ui.setupUi(self)
        self.company_folder = find_company_folder()

        # Path to your Excel file
        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )

        # --- Connect button ---        
        self.ui.nhapairButton.clicked.connect(self.hangdiairsheet)

        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb["THEO DÕI HÀNG ĐI AIR"]

            def get_completer_list(col_idx):
                # col_idx: 1-based index (A=1, B=2, ...)
                return list({
                    str(row[0].value).strip()
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx)
                    if row[0].value
                })

            # lineEdit_3 → column E (5)
            completer3 = QCompleter(get_completer_list(5))
            completer3.setCaseSensitivity(QtCore.Qt.CaseInsensitive)

            # lineEdit_6 → column H (8)
            completer6 = QCompleter(get_completer_list(8))
            completer6.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer6.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_6.setCompleter(completer6)

            # lineEdit_8 → column K (11)
            completer8 = QCompleter(get_completer_list(11))
            completer8.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer8.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_8.setCompleter(completer8)

            # lineEdit_15 → column S (19)
            completer15 = QCompleter(get_completer_list(19))
            completer15.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            completer15.setCompletionMode(QCompleter.PopupCompletion)
            self.ui.lineEdit_15.setCompleter(completer15)

            wb.close()

    def hangdiairsheet(self):
        import openpyxl

        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )

        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["THEO DÕI HÀNG ĐI AIR"]
            next_row = ws.max_row + 1

            ws.cell(row=next_row, column=1, value=self.ui.lineEdit_1.text())    # A
            ws.cell(row=next_row, column=2, value=self.ui.lineEdit_2.text())    # B
            ws.cell(row=next_row, column=5, value=self.ui.lineEdit_3.text())    # E
            ws.cell(row=next_row, column=6, value=self.ui.lineEdit_4.text())    # F
            ws.cell(row=next_row, column=7, value=self.ui.lineEdit_5.text())    # G
            ws.cell(row=next_row, column=8, value=self.ui.lineEdit_6.text())    # H
            ws.cell(row=next_row, column=10, value=self.ui.lineEdit_7.text())   # J
            ws.cell(row=next_row, column=11, value=self.ui.lineEdit_8.text())   # K
            ws.cell(row=next_row, column=12, value=self.ui.lineEdit_9.text())   # L
            ws.cell(row=next_row, column=13, value=self.ui.lineEdit_10.text())  # M
            ws.cell(row=next_row, column=14, value=self.ui.lineEdit_11.text())  # N
            ws.cell(row=next_row, column=15, value=self.ui.lineEdit_12.text())  # O
            ws.cell(row=next_row, column=16, value=self.ui.lineEdit_13.text())  # P
            ws.cell(row=next_row, column=17, value=self.ui.lineEdit_14.text())  # Q
            ws.cell(row=next_row, column=19, value=self.ui.lineEdit_15.text())  # S

            wb.save(excel_path)
            wb.close()
            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã nhập dữ liệu AIR thành công!"))
            for i in range(1, 15):
                getattr(self.ui, f"lineEdit_{i}").setText("")
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không thể ghi file Excel:\n{e}"))

class LLCSampleFormDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_LLCSampleForm()
        self.ui.setupUi(self)
        self.company_folder = find_company_folder()

        # --- Connect button ---        
        self.ui.nhapllcButton.clicked.connect(self.llcsamplesheet)

    def llcsamplesheet(self):
        import openpyxl

        excel_path = os.path.join(
            self.company_folder,
            "THE LACQUER COMPANY - VIETNAM OFFICE",
            "WAREHOUSE - IN OUT",
            "00. INVENTORY REPORT",
            "VIETNAM STOCK LIST UPDATE-APP.xlsx"
        )

        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["LLC SAMPLE"]
            next_row = ws.max_row + 1

            ws.cell(row=next_row, column=1, value=self.ui.lineEdit_1.text())    # A
            ws.cell(row=next_row, column=2, value=self.ui.lineEdit_2.text())    # B
            ws.cell(row=next_row, column=3, value=self.ui.lineEdit_3.text())    # C
            ws.cell(row=next_row, column=4, value=self.ui.lineEdit_4.text())    # D
            ws.cell(row=next_row, column=5, value=self.ui.lineEdit_5.text())    # E
            ws.cell(row=next_row, column=6, value=self.ui.lineEdit_6.text())    # F
            ws.cell(row=next_row, column=7, value=self.ui.lineEdit_7.text())    # G
            ws.cell(row=next_row, column=8, value=self.ui.lineEdit_8.text())    # H
            ws.cell(row=next_row, column=9, value=self.ui.lineEdit_9.text())    # I
            ws.cell(row=next_row, column=10, value=self.ui.lineEdit_10.text())  # J
            ws.cell(row=next_row, column=11, value=self.ui.lineEdit_11.text())  # K
            ws.cell(row=next_row, column=12, value=self.ui.lineEdit_12.text())  # L
            ws.cell(row=next_row, column=14, value=self.ui.lineEdit_13.text())  # N

            wb.save(excel_path)
            wb.close()
            QtWidgets.QMessageBox.information(self, self.tr("Thành công"), self.tr("Đã nhập dữ liệu LLC SAMPLE thành công!"))
            for i in range(1, 13):
                getattr(self.ui, f"lineEdit_{i}").setText("")
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, self.tr("Error"), self.tr(f"Không thể ghi file Excel:\n{e}"))


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = WarehouseForm()
    window.show()
    sys.exit(app.exec_())